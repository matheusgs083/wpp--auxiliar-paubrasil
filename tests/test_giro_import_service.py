from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook

from bot_api.services.giro_import_service import STATUS_NOK, STATUS_ZERO, _load_giro_rows


class GiroImportServiceTests(unittest.TestCase):
    def test_load_rows_accepts_visit_column_and_empty_separators(self) -> None:
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "giro.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Giro"
            sheet.append([None] * 25)
            sheet.append(["Tipo", None, None, None, None, "Litrinho", None, None, None, None, None, None, "Inteira", None, None, None, None, None, None, "Litrao"])
            sheet.append(
                [
                    "Revenda",
                    "Fantasia",
                    "Setor",
                    "Visita",
                    None,
                    "NB",
                    "Filial",
                    "Total",
                    "Real",
                    "Gap",
                    "Giro",
                    None,
                    "NB",
                    "Filial",
                    "Total",
                    "Real",
                    "Gap",
                    "Giro",
                    None,
                    "NB",
                    "Filial",
                    "Total",
                    "Real",
                    "Gap",
                    "Giro",
                ]
            )
            sheet.append(["Sousa", "AABB", 138, "SEG/", None, 12068, 1, 0, 0, 0, "-", None, 12068, 1, 14, 0, 28, "Zero", None, 12068, 1, 3, 0, 6, "Zero"])
            sheet.append([None] * 25)
            sheet.append([None, "ACONCHEGO", 130, "TER/", None, 99458, 1, 80, 0, 160, "Zero", None, 99458, 1, 145, 111, 179, "NOK", None, 99458, 1, 0, 0, 0, "-"])
            workbook.save(path)
            workbook.close()

            sheet_name, rows = _load_giro_rows(path)

        self.assertEqual(sheet_name, "Giro")
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0].revenda, "Sousa")
        self.assertEqual(rows[0].fantasia, "AABB")
        self.assertEqual(rows[0].setor, "138")
        self.assertEqual(rows[0].nb, "12068")
        self.assertEqual(rows[0].filial, "1")
        self.assertEqual(rows[0].gap_inteira, 28)
        self.assertEqual(rows[0].giro_inteira, STATUS_ZERO)
        self.assertEqual(rows[1].revenda, "Sousa")
        self.assertEqual(rows[1].nb, "99458")
        self.assertEqual(rows[1].gap_litrinho, 160)
        self.assertEqual(rows[1].giro_inteira, STATUS_NOK)

    def test_load_rows_accepts_blank_column_between_filial_and_total(self) -> None:
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "giro.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Giro"
            sheet.append(["Tipo", None, None, None, "Litrinho", "Litrinho", "Litrinho", "Litrinho", "Litrinho", "Litrinho", "Litrinho", "Inteira", "Inteira", "Inteira", "Inteira", "Inteira", "Inteira", "Inteira", "Litrao"])
            sheet.append(
                [
                    "Revenda",
                    "Fantasia",
                    "Setor",
                    "Visita",
                    "NB",
                    "Filial",
                    None,
                    "Total",
                    "Real",
                    "Gap",
                    "Giro",
                    "NB",
                    "Filial",
                    None,
                    "Total",
                    "Real",
                    "Gap",
                    "Giro",
                    "NB",
                    "Filial",
                    None,
                    "Total",
                    "Real",
                    "Gap",
                    "Giro",
                ]
            )
            sheet.append(["Sousa", "ACONCHEGO", 130, "TER/", 99458, 1, None, 80, 0, 160, "Zero", 99458, 1, None, 145, 111, 179, "NOK", 99458, 1, None, 0, 0, 0, "-"])
            workbook.save(path)
            workbook.close()

            _sheet_name, rows = _load_giro_rows(path)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].nb, "99458")
        self.assertEqual(rows[0].filial, "1")
        self.assertEqual(rows[0].total_litrinho, 80)
        self.assertEqual(rows[0].gap_litrinho, 160)
        self.assertEqual(rows[0].total_inteira, 145)
        self.assertEqual(rows[0].real_inteira, 111)
        self.assertEqual(rows[0].gap_inteira, 179)

    def test_modern_layout_prefers_revenda_as_filial_when_filial_column_is_generic(self) -> None:
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "giro.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Giro"
            sheet.append(["Tipo", None, None, None, "Litrinho", None, None, None, None, None, None, "Inteira", None, None, None, None, None, None, "Litrao"])
            sheet.append(
                [
                    "Revenda",
                    "Fantasia",
                    "Setor",
                    "Visita",
                    "NB",
                    "Filial",
                    None,
                    "Total",
                    "Real",
                    "Gap",
                    "Giro",
                    "NB",
                    "Filial",
                    None,
                    "Total",
                    "Real",
                    "Gap",
                    "Giro",
                    "NB",
                    "Filial",
                    None,
                    "Total",
                    "Real",
                    "Gap",
                    "Giro",
                ]
            )
            sheet.append(["Sousa", "CLIENTE SOUSA", 125, "SEG/", 310, 1, None, 10, 0, 20, "Zero", 310, 1, None, 0, 0, 0, "-", 310, 1, None, 0, 0, 0, "-"])
            sheet.append(["Patos", "CLIENTE PATOS", 401, "TER/", 310, 1, None, 0, 0, 0, "-", 310, 1, None, 12, 0, 24, "Zero", 310, 1, None, 0, 0, 0, "-"])
            sheet.append(["Sumé", "CLIENTE SUME", 501, "QUA/", 310, 1, None, 0, 0, 0, "-", 310, 1, None, 0, 0, 0, "-", 310, 1, None, 8, 0, 16, "Zero"])
            sheet.append(["Filtros aplicados:\nTipo não está em branco", None, None, None, None, 1, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None])
            workbook.save(path)
            workbook.close()

            _sheet_name, rows = _load_giro_rows(path)

        self.assertEqual([(row.revenda, row.filial, row.nb) for row in rows], [("Sousa", "1", "310"), ("Patos", "3", "310"), ("Sumé", "4", "310")])


if __name__ == "__main__":
    unittest.main()
