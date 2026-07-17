from __future__ import annotations

import csv
import hashlib
import io
import unicodedata
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import psycopg
from psycopg import sql
from psycopg.rows import dict_row

from bot_api.commercial_scope import normalize_numeric_code
from bot_api.services.import_publication import (
    activate_import_batch,
    ensure_dataset_state_table,
    prune_import_batches,
    resolve_effective_import_batch_id,
)


EXPECTED_EXTENSION_SET = {".xlsx", ".xlsm", ".csv"}


@dataclass(frozen=True)
class DRevendasValidationResult:
    dataset_name: str
    source_path: str
    ok: bool
    total_rows: int
    unique_revendas: int
    error_count: int
    warning_count: int
    sample_errors: list[str]
    sample_warnings: list[str]

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)

    def ensure_valid(self) -> None:
        if self.error_count:
            raise ValueError("; ".join(self.sample_errors) if self.sample_errors else "Erros na validacao da tabela de revendas.")


@dataclass(frozen=True)
class DRevendasImportSummary:
    source_path: str
    rows: int
    unique_revendas: int

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class DRevendasRow:
    codigo: str
    nome: str
    source_row_number: int


class DRevendasImportService:
    def __init__(self, database_url: str, schema: str, connect_timeout_seconds: float = 3.0) -> None:
        self.database_url = database_url.strip()
        self.schema = _normalize_schema(schema)
        self.connect_timeout_seconds = max(float(connect_timeout_seconds), 1.0)

    def validate_source(self, source_path: Path) -> DRevendasValidationResult:
        path = source_path.expanduser().resolve()
        errors: list[str] = []
        warnings_list: list[str] = []
        rows: list[DRevendasRow] = []

        if not path.exists():
            errors.append(f"Arquivo nao encontrado: {path}")
        elif path.suffix.lower() not in EXPECTED_EXTENSION_SET:
            errors.append("Formato invalido. Use um arquivo .xlsx, .xlsm ou .csv.")
        else:
            try:
                rows = _load_drevendas_rows(path)
            except Exception as exc:
                errors.append(str(exc))

        if not rows and not errors:
            errors.append("Nao encontrei linhas validas na tabela de revendas.")

        return DRevendasValidationResult(
            dataset_name="drevendas",
            source_path=str(path),
            ok=not errors,
            total_rows=len(rows),
            unique_revendas=len({row.codigo for row in rows}),
            error_count=len(errors),
            warning_count=len(warnings_list),
            sample_errors=errors[:10],
            sample_warnings=warnings_list[:10],
        )

    def summarize_source(self, source_path: Path) -> DRevendasImportSummary:
        validation = self.validate_source(source_path)
        validation.ensure_valid()
        rows = _load_drevendas_rows(source_path.expanduser().resolve())
        return DRevendasImportSummary(
            source_path=str(source_path),
            rows=len(rows),
            unique_revendas=len({row.codigo for row in rows}),
        )

    def import_source(self, source_path: Path, reference_date: date | None = None) -> dict[str, Any]:
        if not self.database_url:
            raise RuntimeError("REPORTS_DATABASE_URL nao configurada.")

        path = source_path.expanduser().resolve()
        validation = self.validate_source(path)
        validation.ensure_valid()
        summary = self.summarize_source(path)
        rows = _load_drevendas_rows(path)
        batch_date = reference_date or datetime.fromtimestamp(path.stat().st_mtime).date()
        source_hash = _sha256(path)

        with self._connect() as conn:
            self._ensure_schema(conn)
            batch_id = self._insert_batch(conn, str(path), batch_date, source_hash, len(rows))
            self._insert_snapshot_rows(conn, rows, batch_id)
            activate_import_batch(conn, self.schema, "drevendas", batch_id)
            self._create_latest_view(conn)
            prune_import_batches(conn, self.schema, "drevendas", keep_last=3)
            conn.commit()

        result = summary.to_dict()
        result.update(
            {
                "batch_id": batch_id,
                "reference_date": batch_date.isoformat(),
                "source_hash": source_hash,
                "schema": self.schema,
                "replaced_previous_batches": False,
                "published_as_active_batch": True,
            }
        )
        return result

    def refresh_latest_view(self) -> dict[str, Any]:
        if not self.database_url:
            raise RuntimeError("REPORTS_DATABASE_URL nao configurada.")

        with self._connect() as conn:
            self._ensure_schema(conn)
            active_batch_id = resolve_effective_import_batch_id(conn, self.schema, "drevendas", activate_if_missing=True)
            self._create_latest_view(conn)
            conn.commit()

        return {
            "ok": True,
            "schema": self.schema,
            "view": f"{self.schema}.drevendas_latest",
            "active_batch_id": active_batch_id,
        }

    def latest_labels(self) -> dict[str, str]:
        if not self.database_url:
            return {}
        try:
            with self._connect() as conn:
                self._ensure_schema(conn)
                self._create_latest_view(conn)
                with conn.cursor(row_factory=dict_row) as cur:
                    cur.execute(
                        sql.SQL("SELECT codigo, nome FROM {}.drevendas_latest ORDER BY codigo::integer, codigo").format(
                            sql.Identifier(self.schema)
                        )
                    )
                    rows = cur.fetchall()
                conn.commit()
        except Exception:
            return {}
        return {str(row["codigo"]): str(row["nome"]) for row in rows if row.get("codigo") and row.get("nome")}

    def _ensure_schema(self, conn: psycopg.Connection[Any]) -> None:
        with conn.cursor() as cur:
            cur.execute(sql.SQL("CREATE SCHEMA IF NOT EXISTS {}").format(sql.Identifier(self.schema)))
            cur.execute(
                sql.SQL(
                    """
                    CREATE TABLE IF NOT EXISTS {}.import_batches (
                        id BIGSERIAL PRIMARY KEY,
                        dataset_name VARCHAR(80) NOT NULL,
                        source_file TEXT NOT NULL,
                        file_hash VARCHAR(64) NOT NULL,
                        reference_date DATE,
                        total_rows INTEGER NOT NULL,
                        imported_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                    """
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    """
                    CREATE TABLE IF NOT EXISTS {}.drevendas_snapshot (
                        batch_id BIGINT NOT NULL REFERENCES {}.import_batches(id) ON DELETE CASCADE,
                        row_number BIGINT NOT NULL,
                        codigo VARCHAR(32) NOT NULL,
                        nome TEXT NOT NULL,
                        imported_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        PRIMARY KEY (batch_id, row_number)
                    )
                    """
                ).format(sql.Identifier(self.schema), sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS drevendas_snapshot_batch_codigo_idx ON {}.drevendas_snapshot (batch_id, codigo)"
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS import_batches_drevendas_dataset_idx ON {}.import_batches (dataset_name, imported_at DESC)"
                ).format(sql.Identifier(self.schema))
            )
            ensure_dataset_state_table(conn, self.schema)

    def _insert_batch(
        self,
        conn: psycopg.Connection[Any],
        source_file: str,
        reference_date: date,
        file_hash: str,
        total_rows: int,
    ) -> int:
        query = sql.SQL(
            """
            INSERT INTO {}.import_batches (dataset_name, source_file, file_hash, reference_date, total_rows)
            VALUES ('drevendas', %s, %s, %s, %s)
            RETURNING id
            """
        ).format(sql.Identifier(self.schema))
        with conn.cursor(row_factory=dict_row) as cur:
            cur.execute(query, (source_file, file_hash, reference_date, total_rows))
            row = cur.fetchone()
        return int(row["id"])

    def _insert_snapshot_rows(self, conn: psycopg.Connection[Any], rows: list[DRevendasRow], batch_id: int) -> None:
        query = sql.SQL(
            """
            INSERT INTO {}.drevendas_snapshot (
                batch_id,
                row_number,
                codigo,
                nome
            )
            VALUES (%s, %s, %s, %s)
            """
        ).format(sql.Identifier(self.schema))
        params = [(batch_id, index, row.codigo, row.nome) for index, row in enumerate(rows, start=1)]
        with conn.cursor() as cur:
            cur.executemany(query, params)

    def _create_latest_view(self, conn: psycopg.Connection[Any]) -> None:
        active_batch_id = resolve_effective_import_batch_id(conn, self.schema, "drevendas", activate_if_missing=True)
        where_clause = sql.SQL("s.batch_id = {}").format(sql.Literal(active_batch_id)) if active_batch_id is not None else sql.SQL("FALSE")
        query = sql.SQL(
            """
            CREATE OR REPLACE VIEW {}.drevendas_latest AS
            SELECT
                s.batch_id,
                s.row_number,
                s.codigo,
                s.nome,
                s.imported_at,
                b.reference_date,
                b.source_file,
                b.file_hash,
                b.imported_at AS batch_imported_at
            FROM {}.drevendas_snapshot s
            JOIN {}.import_batches b ON b.id = s.batch_id
            WHERE {}
            """
        ).format(
            sql.Identifier(self.schema),
            sql.Identifier(self.schema),
            sql.Identifier(self.schema),
            where_clause,
        )
        with conn.cursor() as cur:
            cur.execute(query)

    def _connect(self) -> psycopg.Connection[Any]:
        return psycopg.connect(self.database_url, connect_timeout=self.connect_timeout_seconds)


def _load_drevendas_rows(path: Path) -> list[DRevendasRow]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _load_drevendas_rows_from_workbook(path)
    if suffix == ".csv":
        return _load_drevendas_rows_from_csv(path)
    raise ValueError("Formato invalido. Use um arquivo .xlsx, .xlsm ou .csv.")


def _load_drevendas_rows_from_workbook(path: Path) -> list[DRevendasRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration as exc:
            raise ValueError("Arquivo sem cabecalho.") from exc
        indexes = _header_indexes(list(header or []))
        rows: list[DRevendasRow] = []
        for row_number, row in enumerate(rows_iter, start=2):
            parsed = _parse_row(row, row_number, indexes)
            if parsed is not None:
                rows.append(parsed)
        return rows
    finally:
        workbook.close()


def _load_drevendas_rows_from_csv(path: Path) -> list[DRevendasRow]:
    text = _read_text_with_fallback(path)
    delimiter = _detect_delimiter(text)
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    try:
        header = next(reader)
    except StopIteration as exc:
        raise ValueError("Arquivo sem cabecalho.") from exc
    indexes = _header_indexes(header)
    rows: list[DRevendasRow] = []
    for row_number, row in enumerate(reader, start=2):
        parsed = _parse_row(row, row_number, indexes)
        if parsed is not None:
            rows.append(parsed)
    return rows


def _header_indexes(header: list[Any]) -> tuple[int, int]:
    normalized = [_normalize_header(value) for value in header]
    codigo_index = _first_index(normalized, {"unb", "filial", "revenda", "codigo", "codigorevenda", "cod"})
    nome_index = _first_index(normalized, {"nome", "revenda", "nomerevenda", "descricao", "operacao"})
    if codigo_index is None or nome_index is None or codigo_index == nome_index:
        if len(header) < 2:
            raise ValueError("Arquivo invalido. Informe pelo menos as colunas UNB e NOME.")
        return 0, 1
    return codigo_index, nome_index


def _first_index(values: list[str], candidates: set[str]) -> int | None:
    for index, value in enumerate(values):
        if value in candidates:
            return index
    return None


def _parse_row(row: Any, row_number: int, indexes: tuple[int, int]) -> DRevendasRow | None:
    values = list(row or [])
    codigo_index, nome_index = indexes
    codigo = normalize_numeric_code(str(values[codigo_index] if codigo_index < len(values) else "").strip())
    nome = _clean_text(values[nome_index] if nome_index < len(values) else "")
    if not codigo and not nome:
        return None
    if not codigo or not nome:
        return None
    return DRevendasRow(codigo=codigo, nome=nome, source_row_number=row_number)


def _read_text_with_fallback(path: Path) -> str:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("Nao consegui ler o arquivo com um encoding suportado.")


def _detect_delimiter(text: str) -> str:
    sample = "\n".join(text.splitlines()[:10])
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;")
        return str(dialect.delimiter or ";")
    except Exception:
        return ";"


def _normalize_schema(value: str) -> str:
    cleaned = "".join(char for char in str(value or "").strip() if char.isalnum() or char == "_")
    return cleaned or "reports"


def _normalize_header(value: Any) -> str:
    raw = str(value or "").strip().lower()
    normalized = unicodedata.normalize("NFKD", raw)
    without_accents = "".join(char for char in normalized if not unicodedata.combining(char))
    without_accents = "".join(char for char in without_accents if char.isalnum())
    return without_accents


def _clean_text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fp:
        for chunk in iter(lambda: fp.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
