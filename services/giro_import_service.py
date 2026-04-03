from __future__ import annotations

import hashlib
from dataclasses import asdict, dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
import warnings

from openpyxl import load_workbook
import psycopg
from psycopg import sql
from psycopg.rows import dict_row
from psycopg.types.json import Jsonb

from bot_api.commercial_scope import normalize_numeric_code
from bot_api.services.dsetores_import_service import ensure_dsetores_schema


EXPECTED_EXTENSION_SET = {".xlsx", ".xlsm", ".xls"}
STATUS_OK = "OK"
STATUS_NOK = "NOK"
STATUS_ZERO = "ZERO"
STATUS_EMPTY = "-"


@dataclass(frozen=True)
class GiroValidationResult:
    dataset_name: str
    source_path: str
    ok: bool
    sheet_name: str
    total_rows: int
    error_count: int
    warning_count: int
    sample_errors: list[str]
    sample_warnings: list[str]

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)

    def ensure_valid(self) -> None:
        if self.error_count:
            joined = "; ".join(self.sample_errors) if self.sample_errors else "Erros na validacao do giro."
            raise ValueError(joined)


@dataclass(frozen=True)
class GiroImportSummary:
    source_path: str
    sheet_name: str
    rows: int
    unique_filiais: int
    unique_setores: int
    unique_filial_setores: int
    unique_clientes: int
    attention_clients: int
    zero_clients: int
    monitored_litrinho: int
    monitored_inteira: int
    monitored_litrao: int

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class GiroFlatRow:
    revenda: str
    fantasia: str
    setor: str
    nb: str
    filial: str
    total_litrinho: Decimal
    real_litrinho: Decimal
    gap_litrinho: Decimal
    giro_litrinho: str
    total_inteira: Decimal
    real_inteira: Decimal
    gap_inteira: Decimal
    giro_inteira: str
    total_litrao: Decimal
    real_litrao: Decimal
    gap_litrao: Decimal
    giro_litrao: str
    source_row_number: int

    def to_payload(self) -> dict[str, Any]:
        return {
            "revenda": self.revenda,
            "fantasia": self.fantasia,
            "setor": self.setor,
            "nb": self.nb,
            "filial": self.filial,
            "total_litrinho": _decimal_to_json(self.total_litrinho),
            "real_litrinho": _decimal_to_json(self.real_litrinho),
            "gap_litrinho": _decimal_to_json(self.gap_litrinho),
            "giro_litrinho": self.giro_litrinho,
            "total_inteira": _decimal_to_json(self.total_inteira),
            "real_inteira": _decimal_to_json(self.real_inteira),
            "gap_inteira": _decimal_to_json(self.gap_inteira),
            "giro_inteira": self.giro_inteira,
            "total_litrao": _decimal_to_json(self.total_litrao),
            "real_litrao": _decimal_to_json(self.real_litrao),
            "gap_litrao": _decimal_to_json(self.gap_litrao),
            "giro_litrao": self.giro_litrao,
            "source_row_number": self.source_row_number,
        }


class GiroImportService:
    def __init__(self, database_url: str, schema: str, connect_timeout_seconds: float = 3.0) -> None:
        self.database_url = database_url.strip()
        self.schema = _normalize_schema(schema)
        self.connect_timeout_seconds = max(float(connect_timeout_seconds), 1.0)

    def validate_workbook(self, source_path: Path) -> GiroValidationResult:
        path = source_path.expanduser().resolve()
        errors: list[str] = []
        warnings_list: list[str] = []

        if not path.exists():
            errors.append(f"Arquivo nao encontrado: {path}")
            return GiroValidationResult(
                dataset_name="giro",
                source_path=str(path),
                ok=False,
                sheet_name="",
                total_rows=0,
                error_count=len(errors),
                warning_count=len(warnings_list),
                sample_errors=errors[:10],
                sample_warnings=warnings_list[:10],
            )

        if path.suffix.lower() not in EXPECTED_EXTENSION_SET:
            errors.append("Formato invalido. Use .xlsx, .xlsm ou .xls.")

        try:
            sheet_name, rows = _load_giro_rows(path)
        except Exception as exc:
            errors.append(str(exc))
            sheet_name = ""
            rows = []

        if not rows:
            errors.append("Nao encontrei linhas validas na planilha de giro.")

        for row in rows[:20]:
            if not row.filial:
                errors.append(f"Linha {row.source_row_number}: filial vazia.")
            if not row.nb:
                errors.append(f"Linha {row.source_row_number}: NB vazio.")
            if not row.setor:
                warnings_list.append(f"Linha {row.source_row_number}: setor vazio.")
            if not row.fantasia:
                warnings_list.append(f"Linha {row.source_row_number}: fantasia vazia.")
            if len(errors) >= 10:
                break

        return GiroValidationResult(
            dataset_name="giro",
            source_path=str(path),
            ok=not errors,
            sheet_name=sheet_name,
            total_rows=len(rows),
            error_count=len(errors),
            warning_count=len(warnings_list),
            sample_errors=errors[:10],
            sample_warnings=warnings_list[:10],
        )

    def summarize_workbook(self, source_path: Path) -> GiroImportSummary:
        validation = self.validate_workbook(source_path)
        validation.ensure_valid()
        sheet_name, rows = _load_giro_rows(source_path.expanduser().resolve())

        filiais = {row.filial for row in rows if row.filial}
        setores = {row.setor for row in rows if row.setor}
        filial_setores = {(row.filial, row.setor) for row in rows if row.filial and row.setor}
        clientes = {(row.filial, row.nb) for row in rows if row.filial and row.nb}

        return GiroImportSummary(
            source_path=str(source_path),
            sheet_name=sheet_name,
            rows=len(rows),
            unique_filiais=len(filiais),
            unique_setores=len(setores),
            unique_filial_setores=len(filial_setores),
            unique_clientes=len(clientes),
            attention_clients=sum(1 for row in rows if _row_has_attention(row)),
            zero_clients=sum(1 for row in rows if _row_has_zero(row)),
            monitored_litrinho=sum(1 for row in rows if row.giro_litrinho != STATUS_EMPTY),
            monitored_inteira=sum(1 for row in rows if row.giro_inteira != STATUS_EMPTY),
            monitored_litrao=sum(1 for row in rows if row.giro_litrao != STATUS_EMPTY),
        )

    def import_workbook(self, source_path: Path, reference_date: date | None = None) -> dict[str, Any]:
        if not self.database_url:
            raise RuntimeError("REPORTS_DATABASE_URL nao configurada.")

        path = source_path.expanduser().resolve()
        validation = self.validate_workbook(path)
        validation.ensure_valid()
        summary = self.summarize_workbook(path)
        batch_date = reference_date or datetime.fromtimestamp(path.stat().st_mtime).date()
        file_hash = _sha256(path)
        sheet_name, rows = _load_giro_rows(path)

        with self._connect() as conn:
            ensure_dsetores_schema(conn, self.schema)
            self._ensure_schema(conn)
            self._replace_dataset_contents(conn, dataset_name="giro")
            batch_id = self._insert_batch(conn, str(path), batch_date, file_hash, len(rows))
            self._insert_snapshot_rows(conn, rows, batch_id)
            self._hydrate_scope_columns(conn, batch_id)
            self._create_latest_view(conn)
            conn.commit()

        result = summary.to_dict()
        result.update(
            {
                "batch_id": batch_id,
                "reference_date": batch_date.isoformat(),
                "file_hash": file_hash,
                "schema": self.schema,
                "sheet_name": sheet_name,
                "replaced_previous_batches": True,
            }
        )
        return result

    def _ensure_schema(self, conn: psycopg.Connection[Any]) -> None:
        with conn.cursor() as cur:
            cur.execute(sql.SQL("CREATE SCHEMA IF NOT EXISTS {}").format(sql.Identifier(self.schema)))
            cur.execute(
                sql.SQL(
                    """
                    CREATE TABLE IF NOT EXISTS {}.giro_snapshot (
                        batch_id BIGINT NOT NULL REFERENCES {}.import_batches(id) ON DELETE CASCADE,
                        row_number INTEGER NOT NULL,
                        revenda TEXT,
                        fantasia TEXT,
                        setor VARCHAR(32) NOT NULL DEFAULT '',
                        nb VARCHAR(32) NOT NULL DEFAULT '',
                        filial VARCHAR(16) NOT NULL DEFAULT '',
                        filial_setor_key TEXT NOT NULL DEFAULT '',
                        filial_gv_key TEXT NOT NULL DEFAULT '',
                        filial_dc_key TEXT NOT NULL DEFAULT '',
                        total_litrinho NUMERIC(18,2) NOT NULL DEFAULT 0,
                        real_litrinho NUMERIC(18,2) NOT NULL DEFAULT 0,
                        gap_litrinho NUMERIC(18,2) NOT NULL DEFAULT 0,
                        giro_litrinho VARCHAR(16) NOT NULL DEFAULT '-',
                        total_inteira NUMERIC(18,2) NOT NULL DEFAULT 0,
                        real_inteira NUMERIC(18,2) NOT NULL DEFAULT 0,
                        gap_inteira NUMERIC(18,2) NOT NULL DEFAULT 0,
                        giro_inteira VARCHAR(16) NOT NULL DEFAULT '-',
                        total_litrao NUMERIC(18,2) NOT NULL DEFAULT 0,
                        real_litrao NUMERIC(18,2) NOT NULL DEFAULT 0,
                        gap_litrao NUMERIC(18,2) NOT NULL DEFAULT 0,
                        giro_litrao VARCHAR(16) NOT NULL DEFAULT '-',
                        payload JSONB NOT NULL,
                        imported_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        PRIMARY KEY (batch_id, row_number),
                        UNIQUE (batch_id, filial, nb)
                    )
                    """
                ).format(sql.Identifier(self.schema), sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL("CREATE INDEX IF NOT EXISTS giro_snapshot_batch_idx ON {}.giro_snapshot (batch_id)").format(
                    sql.Identifier(self.schema)
                )
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS giro_snapshot_batch_filial_nb_idx ON {}.giro_snapshot (batch_id, filial, nb)"
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS giro_snapshot_batch_filial_setor_key_idx ON {}.giro_snapshot (batch_id, filial_setor_key)"
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS giro_snapshot_batch_filial_gv_key_idx ON {}.giro_snapshot (batch_id, filial_gv_key)"
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS giro_snapshot_batch_filial_dc_key_idx ON {}.giro_snapshot (batch_id, filial_dc_key)"
                ).format(sql.Identifier(self.schema))
            )

    def _insert_batch(
        self,
        conn: psycopg.Connection[Any],
        source_label: str,
        reference_date: date,
        file_hash: str,
        total_rows: int,
    ) -> int:
        query = sql.SQL(
            """
            INSERT INTO {}.import_batches (dataset_name, source_file, file_hash, reference_date, total_rows)
            VALUES (%s, %s, %s, %s, %s)
            RETURNING id
            """
        ).format(sql.Identifier(self.schema))
        with conn.cursor(row_factory=dict_row) as cur:
            cur.execute(query, ("giro", source_label, file_hash, reference_date, total_rows))
            row = cur.fetchone()
        return int(row["id"])

    def _insert_snapshot_rows(self, conn: psycopg.Connection[Any], rows: list[GiroFlatRow], batch_id: int) -> None:
        query = sql.SQL(
            """
            INSERT INTO {}.giro_snapshot (
                batch_id,
                row_number,
                revenda,
                fantasia,
                setor,
                nb,
                filial,
                filial_setor_key,
                total_litrinho,
                real_litrinho,
                gap_litrinho,
                giro_litrinho,
                total_inteira,
                real_inteira,
                gap_inteira,
                giro_inteira,
                total_litrao,
                real_litrao,
                gap_litrao,
                giro_litrao,
                payload
            )
            VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
            )
            """
        ).format(sql.Identifier(self.schema))
        payload_batch: list[tuple[Any, ...]] = []
        with conn.cursor() as cur:
            for row_number, row in enumerate(rows, start=1):
                payload_batch.append(
                    (
                        batch_id,
                        row_number,
                        row.revenda,
                        row.fantasia,
                        row.setor,
                        row.nb,
                        row.filial,
                        _compose_scope_key(row.filial, row.setor),
                        row.total_litrinho,
                        row.real_litrinho,
                        row.gap_litrinho,
                        row.giro_litrinho,
                        row.total_inteira,
                        row.real_inteira,
                        row.gap_inteira,
                        row.giro_inteira,
                        row.total_litrao,
                        row.real_litrao,
                        row.gap_litrao,
                        row.giro_litrao,
                        Jsonb(row.to_payload()),
                    )
                )
                if len(payload_batch) >= 500:
                    cur.executemany(query, payload_batch)
                    payload_batch.clear()
            if payload_batch:
                cur.executemany(query, payload_batch)

    def _hydrate_scope_columns(self, conn: psycopg.Connection[Any], batch_id: int) -> None:
        query = sql.SQL(
            """
            UPDATE {schema}.giro_snapshot AS g
            SET
                filial_gv_key = COALESCE(ds.filial_gv_key, ''),
                filial_dc_key = COALESCE(ds.filial_dc_key, '')
            FROM {schema}.dsetores_latest AS ds
            WHERE g.batch_id = %s
              AND ds.filial_setor_key = g.filial_setor_key
            """
        ).format(schema=sql.Identifier(self.schema))
        with conn.cursor() as cur:
            cur.execute(query, (batch_id,))

    def _create_latest_view(self, conn: psycopg.Connection[Any]) -> None:
        query = sql.SQL(
            """
            CREATE OR REPLACE VIEW {}.giro_latest AS
            WITH latest_batch AS (
                SELECT id
                FROM {}.import_batches
                WHERE dataset_name = 'giro'
                ORDER BY imported_at DESC, id DESC
                LIMIT 1
            )
            SELECT
                g.batch_id,
                g.row_number,
                g.revenda,
                g.fantasia,
                g.setor,
                g.nb,
                g.filial,
                g.filial_setor_key,
                g.filial_gv_key,
                g.filial_dc_key,
                g.total_litrinho,
                g.real_litrinho,
                g.gap_litrinho,
                g.giro_litrinho,
                g.total_inteira,
                g.real_inteira,
                g.gap_inteira,
                g.giro_inteira,
                g.total_litrao,
                g.real_litrao,
                g.gap_litrao,
                g.giro_litrao,
                g.payload,
                g.imported_at,
                b.reference_date,
                b.source_file,
                b.file_hash,
                b.imported_at AS batch_imported_at
            FROM {}.giro_snapshot g
            JOIN {}.import_batches b ON b.id = g.batch_id
            JOIN latest_batch lb ON lb.id = g.batch_id
            """
        ).format(
            sql.Identifier(self.schema),
            sql.Identifier(self.schema),
            sql.Identifier(self.schema),
            sql.Identifier(self.schema),
        )
        with conn.cursor() as cur:
            cur.execute(query)

    def _replace_dataset_contents(self, conn: psycopg.Connection[Any], dataset_name: str) -> None:
        query = sql.SQL("DELETE FROM {}.import_batches WHERE dataset_name = %s").format(sql.Identifier(self.schema))
        with conn.cursor() as cur:
            cur.execute(query, (dataset_name,))

    def _connect(self) -> psycopg.Connection[Any]:
        return psycopg.connect(
            self.database_url,
            autocommit=False,
            connect_timeout=int(self.connect_timeout_seconds),
        )


def _load_giro_rows(source_path: Path) -> tuple[str, list[GiroFlatRow]]:
    path = source_path.expanduser().resolve()
    rows: list[GiroFlatRow] = []
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    current_revenda = ""
    for row_index, row_values in enumerate(worksheet.iter_rows(min_row=3, values_only=True), start=3):
        row = list(row_values)
        if len(row) < 24:
            row.extend([None] * (24 - len(row)))

        revenda_candidate = _clean_text(row[0])
        if revenda_candidate:
            current_revenda = revenda_candidate

        fantasia = _clean_text(row[1])
        setor = normalize_numeric_code(row[2])
        if not fantasia and not setor:
            continue

        nb = normalize_numeric_code(row[3] or row[10] or row[17])
        filial = normalize_numeric_code(row[4] or row[11] or row[18])
        rows.append(
            GiroFlatRow(
                revenda=current_revenda,
                fantasia=fantasia,
                setor=setor,
                nb=nb,
                filial=filial,
                total_litrinho=_to_decimal(row[6]),
                real_litrinho=_to_decimal(row[7]),
                gap_litrinho=_to_decimal(row[8]),
                giro_litrinho=_normalize_giro_status(row[9]),
                total_inteira=_to_decimal(row[13]),
                real_inteira=_to_decimal(row[14]),
                gap_inteira=_to_decimal(row[15]),
                giro_inteira=_normalize_giro_status(row[16]),
                total_litrao=_to_decimal(row[20]),
                real_litrao=_to_decimal(row[21]),
                gap_litrao=_to_decimal(row[22]),
                giro_litrao=_normalize_giro_status(row[23]),
                source_row_number=row_index,
            )
        )
    return worksheet.title, rows


def _normalize_schema(schema: str) -> str:
    normalized = str(schema or "").strip()
    return normalized or "reports"


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _to_decimal(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    raw = str(value).strip().replace(".", "").replace(",", ".")
    if not raw:
        return Decimal("0")
    try:
        return Decimal(raw)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _normalize_giro_status(value: Any) -> str:
    normalized = _clean_text(value).upper()
    if normalized == "OK":
        return STATUS_OK
    if normalized == "NOK":
        return STATUS_NOK
    if normalized == "ZERO":
        return STATUS_ZERO
    return STATUS_EMPTY


def _row_has_attention(row: GiroFlatRow) -> bool:
    return any(status in {STATUS_NOK, STATUS_ZERO} for status in (row.giro_litrinho, row.giro_inteira, row.giro_litrao))


def _row_has_zero(row: GiroFlatRow) -> bool:
    return any(status == STATUS_ZERO for status in (row.giro_litrinho, row.giro_inteira, row.giro_litrao))


def _compose_scope_key(filial: str, code: str) -> str:
    if not filial or not code:
        return ""
    return f"{filial}_{code}"


def _sha256(file_path: Path) -> str:
    digest = hashlib.sha256()
    with file_path.open("rb") as file_pointer:
        for chunk in iter(lambda: file_pointer.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _decimal_to_json(value: Decimal) -> float:
    return float(value)
