from __future__ import annotations

import csv
import hashlib
import io
import unicodedata
import warnings
from dataclasses import asdict, dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import psycopg
from psycopg import sql
from psycopg.rows import dict_row
from psycopg.types.json import Jsonb

from bot_api.commercial_scope import normalize_numeric_code
from bot_api.services.import_publication import (
    activate_import_batch,
    ensure_dataset_state_table,
    prune_import_batches,
    resolve_effective_import_batch_id,
)


EXPECTED_EXTENSION_SET = {".xlsx", ".xlsm", ".csv"}
EXPECTED_HEADERS = {"codigo", "produto", "und", "asr", "sub", "frio", "ttc"}
HEADER_ALIASES = {
    "codigo": "codigo",
    "cod": "codigo",
    "produto": "produto",
    "descricao": "produto",
    "descricaoproduto": "produto",
    "und": "und",
    "unidade": "und",
    "unidades": "und",
    "asr": "asr",
    "sub": "sub",
    "frio": "frio",
    "ttc": "ttc",
}


@dataclass(frozen=True)
class DPrecosValidationResult:
    dataset_name: str
    source_path: str
    ok: bool
    total_rows: int
    unique_codigos: int
    unique_produtos: int
    error_count: int
    warning_count: int
    sample_errors: list[str]
    sample_warnings: list[str]

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)

    def ensure_valid(self) -> None:
        if self.error_count:
            raise ValueError("; ".join(self.sample_errors) if self.sample_errors else "Erros na validacao da tabela de precos.")


@dataclass(frozen=True)
class DPrecosImportSummary:
    source_path: str
    rows: int
    unique_codigos: int
    unique_produtos: int
    total_asr: str
    total_sub: str
    total_frio: str
    total_ttc: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class DPrecosRow:
    codigo: str
    produto: str
    und: Decimal
    asr: Decimal
    sub: Decimal
    frio: Decimal
    ttc: Decimal
    source_row_number: int

    def to_payload(self) -> dict[str, Any]:
        return {
            "codigo": self.codigo,
            "produto": self.produto,
            "und": _decimal_to_json(self.und),
            "asr": _decimal_to_json(self.asr),
            "sub": _decimal_to_json(self.sub),
            "frio": _decimal_to_json(self.frio),
            "ttc": _decimal_to_json(self.ttc),
            "source_row_number": self.source_row_number,
        }


class DPrecosImportService:
    def __init__(self, database_url: str, schema: str, connect_timeout_seconds: float = 3.0) -> None:
        self.database_url = database_url.strip()
        self.schema = _normalize_schema(schema)
        self.connect_timeout_seconds = max(float(connect_timeout_seconds), 1.0)

    def validate_source(self, source_path: Path) -> DPrecosValidationResult:
        path = source_path.expanduser().resolve()
        errors: list[str] = []
        warnings_list: list[str] = []
        rows: list[DPrecosRow] = []

        if not path.exists():
            errors.append(f"Arquivo nao encontrado: {path}")
        elif path.suffix.lower() not in EXPECTED_EXTENSION_SET:
            errors.append("Formato invalido. Use um arquivo .xlsx, .xlsm ou .csv.")
        else:
            try:
                rows = _load_dprecos_rows(path)
            except Exception as exc:
                errors.append(str(exc))

        if not rows and not errors:
            errors.append("Nao encontrei linhas validas na tabela de precos.")

        return DPrecosValidationResult(
            dataset_name="dprecos",
            source_path=str(path),
            ok=not errors,
            total_rows=len(rows),
            unique_codigos=len({row.codigo for row in rows}),
            unique_produtos=len({row.produto for row in rows}),
            error_count=len(errors),
            warning_count=len(warnings_list),
            sample_errors=errors[:10],
            sample_warnings=warnings_list[:10],
        )

    def summarize_source(self, source_path: Path) -> DPrecosImportSummary:
        validation = self.validate_source(source_path)
        validation.ensure_valid()
        rows = _load_dprecos_rows(source_path.expanduser().resolve())
        return DPrecosImportSummary(
            source_path=str(source_path),
            rows=len(rows),
            unique_codigos=len({row.codigo for row in rows}),
            unique_produtos=len({row.produto for row in rows}),
            total_asr=_format_decimal_string(sum((row.asr for row in rows), Decimal("0"))),
            total_sub=_format_decimal_string(sum((row.sub for row in rows), Decimal("0"))),
            total_frio=_format_decimal_string(sum((row.frio for row in rows), Decimal("0"))),
            total_ttc=_format_decimal_string(sum((row.ttc for row in rows), Decimal("0"))),
        )

    def import_source(self, source_path: Path, reference_date: date | None = None) -> dict[str, Any]:
        if not self.database_url:
            raise RuntimeError("REPORTS_DATABASE_URL nao configurada.")

        path = source_path.expanduser().resolve()
        validation = self.validate_source(path)
        validation.ensure_valid()
        summary = self.summarize_source(path)
        rows = _load_dprecos_rows(path)
        batch_date = reference_date or datetime.fromtimestamp(path.stat().st_mtime).date()
        source_hash = _sha256(path)

        with self._connect() as conn:
            self._ensure_schema(conn)
            batch_id = self._insert_batch(conn, str(path), batch_date, source_hash, len(rows))
            self._insert_snapshot_rows(conn, rows, batch_id)
            activate_import_batch(conn, self.schema, "dprecos", batch_id)
            self._create_latest_view(conn)
            prune_import_batches(conn, self.schema, "dprecos", keep_last=3)
            conn.commit()

        result = summary.to_dict()
        result.update(
            {
                "batch_id": batch_id,
                "reference_date": batch_date.isoformat(),
                "source_hash": source_hash,
                "schema": self.schema,
                "replaced_previous_batches": False,
                "published_as_active_batch": True,
            }
        )
        return result

    def refresh_latest_view(self) -> dict[str, Any]:
        if not self.database_url:
            raise RuntimeError("REPORTS_DATABASE_URL nao configurada.")

        with self._connect() as conn:
            self._ensure_schema(conn)
            active_batch_id = resolve_effective_import_batch_id(conn, self.schema, "dprecos", activate_if_missing=True)
            self._create_latest_view(conn)
            conn.commit()

        return {
            "ok": True,
            "schema": self.schema,
            "view": f"{self.schema}.dprecos_latest",
            "active_batch_id": active_batch_id,
        }

    def _ensure_schema(self, conn: psycopg.Connection[Any]) -> None:
        with conn.cursor() as cur:
            cur.execute(sql.SQL("CREATE SCHEMA IF NOT EXISTS {}").format(sql.Identifier(self.schema)))
            cur.execute(
                sql.SQL(
                    """
                    CREATE TABLE IF NOT EXISTS {}.import_batches (
                        id BIGSERIAL PRIMARY KEY,
                        dataset_name VARCHAR(80) NOT NULL,
                        source_file TEXT NOT NULL,
                        file_hash VARCHAR(64) NOT NULL,
                        reference_date DATE,
                        total_rows INTEGER NOT NULL,
                        imported_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                    """
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    """
                    CREATE TABLE IF NOT EXISTS {}.dprecos_snapshot (
                        batch_id BIGINT NOT NULL REFERENCES {}.import_batches(id) ON DELETE CASCADE,
                        row_number BIGINT NOT NULL,
                        codigo VARCHAR(32) NOT NULL,
                        produto TEXT NOT NULL,
                        und NUMERIC(18, 3) NOT NULL,
                        asr NUMERIC(18, 4) NOT NULL,
                        sub NUMERIC(18, 4) NOT NULL,
                        frio NUMERIC(18, 4) NOT NULL,
                        ttc NUMERIC(18, 4) NOT NULL,
                        payload JSONB NOT NULL,
                        imported_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        PRIMARY KEY (batch_id, row_number)
                    )
                    """
                ).format(sql.Identifier(self.schema), sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS dprecos_snapshot_batch_codigo_idx ON {}.dprecos_snapshot (batch_id, codigo)"
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS dprecos_snapshot_batch_produto_idx ON {}.dprecos_snapshot (batch_id, produto)"
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS import_batches_dprecos_dataset_idx ON {}.import_batches (dataset_name, imported_at DESC)"
                ).format(sql.Identifier(self.schema))
            )
            ensure_dataset_state_table(conn, self.schema)

    def _insert_batch(
        self,
        conn: psycopg.Connection[Any],
        source_file: str,
        reference_date: date,
        file_hash: str,
        total_rows: int,
    ) -> int:
        query = sql.SQL(
            """
            INSERT INTO {}.import_batches (dataset_name, source_file, file_hash, reference_date, total_rows)
            VALUES ('dprecos', %s, %s, %s, %s)
            RETURNING id
            """
        ).format(sql.Identifier(self.schema))
        with conn.cursor(row_factory=dict_row) as cur:
            cur.execute(query, (source_file, file_hash, reference_date, total_rows))
            row = cur.fetchone()
        return int(row["id"])

    def _insert_snapshot_rows(
        self,
        conn: psycopg.Connection[Any],
        rows: list[DPrecosRow],
        batch_id: int,
    ) -> None:
        query = sql.SQL(
            """
            INSERT INTO {}.dprecos_snapshot (
                batch_id,
                row_number,
                codigo,
                produto,
                und,
                asr,
                sub,
                frio,
                ttc,
                payload
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
        ).format(sql.Identifier(self.schema))
        params = [
            (
                batch_id,
                index,
                row.codigo,
                row.produto,
                row.und,
                row.asr,
                row.sub,
                row.frio,
                row.ttc,
                Jsonb(row.to_payload()),
            )
            for index, row in enumerate(rows, start=1)
        ]
        with conn.cursor() as cur:
            cur.executemany(query, params)

    def _create_latest_view(self, conn: psycopg.Connection[Any]) -> None:
        active_batch_id = resolve_effective_import_batch_id(conn, self.schema, "dprecos", activate_if_missing=True)
        where_clause = sql.SQL("s.batch_id = {}").format(sql.Literal(active_batch_id)) if active_batch_id is not None else sql.SQL("FALSE")
        query = sql.SQL(
            """
            CREATE OR REPLACE VIEW {}.dprecos_latest AS
            SELECT
                s.batch_id,
                s.row_number,
                s.codigo,
                s.produto,
                s.und,
                s.asr,
                s.sub,
                s.frio,
                s.ttc,
                s.payload,
                s.imported_at,
                b.reference_date,
                b.source_file,
                b.file_hash,
                b.imported_at AS batch_imported_at
            FROM {}.dprecos_snapshot s
            JOIN {}.import_batches b ON b.id = s.batch_id
            WHERE {}
            """
        ).format(
            sql.Identifier(self.schema),
            sql.Identifier(self.schema),
            sql.Identifier(self.schema),
            where_clause,
        )
        with conn.cursor() as cur:
            cur.execute(query)

    def _connect(self) -> psycopg.Connection[Any]:
        return psycopg.connect(self.database_url, connect_timeout=self.connect_timeout_seconds)


def _load_dprecos_rows(path: Path) -> list[DPrecosRow]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _load_dprecos_rows_from_workbook(path)
    return _load_dprecos_rows_from_csv(path)


def _load_dprecos_rows_from_csv(path: Path) -> list[DPrecosRow]:
    text = _read_text_with_fallback(path)
    delimiter = _detect_delimiter(text)
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    if reader.fieldnames is None:
        raise ValueError("Arquivo sem cabecalho.")

    header_map = _build_header_map([str(header or "").strip() for header in reader.fieldnames])
    normalized_rows = [
        {str(key or "").strip(): str(value or "").strip() for key, value in raw_row.items() if key is not None}
        for raw_row in reader
    ]
    return _build_dprecos_rows_from_mapping_rows(normalized_rows, header_map=header_map, row_number_offset=2)


def _load_dprecos_rows_from_workbook(path: Path) -> list[DPrecosRow]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_values is None:
            raise ValueError("Planilha sem cabecalho.")

        header_row = [str(value or "").strip() for value in header_values]
        header_map = _build_header_map(header_row)
        rows_as_mapping: list[dict[str, str]] = []
        for row_values in worksheet.iter_rows(min_row=2, values_only=True):
            row_map: dict[str, str] = {}
            for index, header in enumerate(header_row):
                if not header:
                    continue
                value = row_values[index] if index < len(row_values) else None
                row_map[header] = _cell_to_text(value)
            rows_as_mapping.append(row_map)
        return _build_dprecos_rows_from_mapping_rows(rows_as_mapping, header_map=header_map, row_number_offset=2)
    finally:
        workbook.close()


def _build_header_map(headers: list[str]) -> dict[str, str]:
    header_map: dict[str, str] = {}
    for actual in headers:
        normalized = _normalize_header(actual)
        canonical = HEADER_ALIASES.get(normalized)
        if canonical and canonical not in header_map:
            header_map[canonical] = actual

    missing_headers = sorted(EXPECTED_HEADERS - set(header_map))
    if missing_headers:
        raise ValueError(f"Arquivo invalido. Colunas obrigatorias ausentes: {', '.join(missing_headers)}")
    return header_map


def _build_dprecos_rows_from_mapping_rows(
    raw_rows: list[dict[str, str]],
    *,
    header_map: dict[str, str],
    row_number_offset: int,
) -> list[DPrecosRow]:
    rows: list[DPrecosRow] = []
    for index, row in enumerate(raw_rows, start=row_number_offset):
        if not any(str(value or "").strip() for value in row.values()):
            continue
        codigo = normalize_numeric_code(row.get(header_map["codigo"], ""))
        produto = _clean_product_text(row.get(header_map["produto"], ""))
        if not codigo or not produto:
            continue
        rows.append(
            DPrecosRow(
                codigo=codigo,
                produto=produto,
                und=_parse_decimal_value(row.get(header_map["und"], "")),
                asr=_parse_decimal_value(row.get(header_map["asr"], "")),
                sub=_parse_decimal_value(row.get(header_map["sub"], "")),
                frio=_parse_decimal_value(row.get(header_map["frio"], "")),
                ttc=_parse_decimal_value(row.get(header_map["ttc"], "")),
                source_row_number=index,
            )
        )
    rows.sort(key=lambda item: (_sort_numeric(item.codigo), item.produto, item.source_row_number))
    return rows


def _read_text_with_fallback(path: Path) -> str:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("Nao consegui ler o arquivo com um encoding suportado.")


def _detect_delimiter(text: str) -> str:
    sample = "\n".join(text.splitlines()[:10])
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;")
        return str(dialect.delimiter or ",")
    except Exception:
        return ","


def _normalize_header(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_only = "".join(char for char in normalized if not unicodedata.combining(char))
    return "".join(char for char in ascii_only.lower() if char.isalnum())


def _normalize_schema(value: str) -> str:
    cleaned = "".join(char for char in str(value or "").strip() if char.isalnum() or char == "_")
    return cleaned or "reports"


def _clean_product_text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def _parse_decimal_value(value: Any) -> Decimal:
    cleaned = str(value or "").strip()
    if not cleaned:
        return Decimal("0")
    normalized = cleaned.replace("R$", "").replace(" ", "")
    if "," in normalized and "." in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")
    elif "," in normalized:
        normalized = normalized.replace(",", ".")
    try:
        return Decimal(normalized)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, Decimal)):
        return str(value)
    if isinstance(value, float):
        text = f"{value:.10f}".rstrip("0").rstrip(".")
        return text or "0"
    return str(value).strip()


def _format_decimal_string(value: Decimal) -> str:
    return f"{value:.2f}".replace(".", ",")


def _decimal_to_json(value: Decimal) -> str:
    return f"{value:.4f}".rstrip("0").rstrip(".") if value else "0"


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fp:
        for chunk in iter(lambda: fp.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _sort_numeric(value: str) -> tuple[int, str]:
    cleaned = normalize_numeric_code(value)
    if cleaned:
        try:
            return (0, f"{int(cleaned):09d}")
        except ValueError:
            return (1, cleaned)
    return (2, str(value or "").strip())
