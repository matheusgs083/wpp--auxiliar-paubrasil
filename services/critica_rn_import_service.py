from __future__ import annotations

import csv
import hashlib
import io
import unicodedata
import warnings
from dataclasses import asdict, dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import psycopg
from psycopg import sql
from psycopg.rows import dict_row
from psycopg.types.json import Jsonb

from bot_api.commercial_scope import normalize_numeric_code
from bot_api.services.import_publication import (
    activate_import_batch,
    ensure_dataset_state_table,
    prune_import_batches,
    resolve_effective_import_batch_id,
)


EXPECTED_EXTENSION_SET = {".xlsx", ".xlsm", ".csv"}
EXPECTED_HEADERS = {
    "UNB",
    "Pedido",
    "Data Pedido",
    "Cod. PDV",
    "Nome PDV",
    "Setor",
    "Produto",
    "Quantidade",
    "Unid. Venda",
    "Preco Unitario",
}
HEADER_ALIASES = {
    "unb": "UNB",
    "pedido": "Pedido",
    "datapedido": "Data Pedido",
    "operacao": "Operacao",
    "codpdv": "Cod. PDV",
    "codigopdv": "Cod. PDV",
    "nomepdv": "Nome PDV",
    "setor": "Setor",
    "statuspedido": "Status Pedido",
    "totalpedido": "Total Pedido",
    "totalcliente": "Total Cliente",
    "critica1": "Critica 1",
    "critica2": "Critica 2",
    "critica3": "Critica 3",
    "critica4": "Critica 4",
    "critica5": "Critica 5",
    "critica6": "Critica 6",
    "criticarejeicao": "Critica Rejeicao",
    "produto": "Produto",
    "quantidade": "Quantidade",
    "unidvenda": "Unid. Venda",
    "precounitario": "Preco Unitario",
    "precosadf": "Preco S/ ADF",
    "minimopolitica": "Minimo Politica",
    "tipomovimento": "Tipo Movimento",
    "codigogv": "Codigo GV",
    "codigopgv": "Codigo PGV",
}


@dataclass(frozen=True)
class CriticaRnValidationResult:
    dataset_name: str
    source_path: str
    ok: bool
    total_rows: int
    unique_pedidos: int
    unique_unb_setores: int
    unique_produtos: int
    duplicate_pedido_produto_keys: int
    rows_with_critica: int
    error_count: int
    warning_count: int
    sample_errors: list[str]
    sample_warnings: list[str]

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)

    def ensure_valid(self) -> None:
        if self.error_count:
            raise ValueError("; ".join(self.sample_errors) if self.sample_errors else "Erros na validacao da critica RN.")


@dataclass(frozen=True)
class CriticaRnImportSummary:
    source_path: str
    rows: int
    unique_pedidos: int
    unique_unb_setores: int
    unique_produtos: int
    duplicate_pedido_produto_keys: int
    rows_with_critica: int

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class CriticaRnRow:
    filial: str
    pedido: str
    data_pedido: date | None
    operacao: str
    cod_pdv: str
    nome_pdv: str
    setor: str
    filial_setor_key: str
    status_pedido: str
    total_pedido: Decimal
    total_cliente: Decimal
    critica_text: str
    produto_codigo: str
    produto_key: str
    quantidade: Decimal
    unid_venda: str
    preco_unitario: Decimal
    preco_sem_adf: Decimal
    minimo_politica: Decimal
    tipo_movimento: str
    codigo_gv: str
    codigo_pgv: str
    source_row_number: int
    payload: dict[str, Any]


class CriticaRnImportService:
    def __init__(self, database_url: str, schema: str, connect_timeout_seconds: float = 3.0) -> None:
        self.database_url = database_url.strip()
        self.schema = _normalize_schema(schema)
        self.connect_timeout_seconds = max(float(connect_timeout_seconds), 1.0)

    def validate_source(self, source_path: Path) -> CriticaRnValidationResult:
        path = source_path.expanduser().resolve()
        errors: list[str] = []
        warnings_list: list[str] = []
        rows: list[CriticaRnRow] = []

        if not path.exists():
            errors.append(f"Arquivo nao encontrado: {path}")
        elif path.suffix.lower() not in EXPECTED_EXTENSION_SET:
            errors.append("Formato invalido. Use um arquivo .xlsx, .xlsm ou .csv.")
        else:
            try:
                rows = _load_critica_rn_rows(path)
            except Exception as exc:
                errors.append(str(exc))

        if not rows and not errors:
            errors.append("Nao encontrei linhas validas na planilha de critica RN.")

        stats = _summary_stats(rows)
        return CriticaRnValidationResult(
            dataset_name="critica_rn",
            source_path=str(path),
            ok=not errors,
            total_rows=len(rows),
            unique_pedidos=stats["unique_pedidos"],
            unique_unb_setores=stats["unique_unb_setores"],
            unique_produtos=stats["unique_produtos"],
            duplicate_pedido_produto_keys=stats["duplicate_pedido_produto_keys"],
            rows_with_critica=stats["rows_with_critica"],
            error_count=len(errors),
            warning_count=len(warnings_list),
            sample_errors=errors[:10],
            sample_warnings=warnings_list[:10],
        )

    def summarize_source(self, source_path: Path) -> CriticaRnImportSummary:
        validation = self.validate_source(source_path)
        validation.ensure_valid()
        rows = _load_critica_rn_rows(source_path.expanduser().resolve())
        stats = _summary_stats(rows)
        return CriticaRnImportSummary(
            source_path=str(source_path),
            rows=len(rows),
            unique_pedidos=stats["unique_pedidos"],
            unique_unb_setores=stats["unique_unb_setores"],
            unique_produtos=stats["unique_produtos"],
            duplicate_pedido_produto_keys=stats["duplicate_pedido_produto_keys"],
            rows_with_critica=stats["rows_with_critica"],
        )

    def import_source(self, source_path: Path, reference_date: date | None = None) -> dict[str, Any]:
        if not self.database_url:
            raise RuntimeError("REPORTS_DATABASE_URL nao configurada.")

        path = source_path.expanduser().resolve()
        validation = self.validate_source(path)
        validation.ensure_valid()
        summary = self.summarize_source(path)
        rows = _load_critica_rn_rows(path)
        batch_date = reference_date or datetime.fromtimestamp(path.stat().st_mtime).date()
        source_hash = _sha256(path)

        with self._connect() as conn:
            self._ensure_schema(conn)
            batch_id = self._insert_batch(conn, str(path), batch_date, source_hash, len(rows))
            self._insert_snapshot_rows(conn, rows, batch_id)
            self._hydrate_scope_columns(conn, batch_id)
            activate_import_batch(conn, self.schema, "critica_rn", batch_id)
            self._create_latest_view(conn)
            prune_import_batches(conn, self.schema, "critica_rn", keep_last=3)
            conn.commit()

        result = summary.to_dict()
        result.update(
            {
                "batch_id": batch_id,
                "reference_date": batch_date.isoformat(),
                "source_hash": source_hash,
                "schema": self.schema,
                "replaced_previous_batches": False,
                "published_as_active_batch": True,
            }
        )
        return result

    def refresh_latest_view(self) -> dict[str, Any]:
        if not self.database_url:
            raise RuntimeError("REPORTS_DATABASE_URL nao configurada.")

        with self._connect() as conn:
            self._ensure_schema(conn)
            active_batch_id = resolve_effective_import_batch_id(conn, self.schema, "critica_rn", activate_if_missing=True)
            if active_batch_id is not None:
                self._hydrate_scope_columns(conn, active_batch_id)
            self._create_latest_view(conn)
            conn.commit()

        return {
            "ok": True,
            "schema": self.schema,
            "view": f"{self.schema}.critica_rn_latest",
            "active_batch_id": active_batch_id,
        }

    def _ensure_schema(self, conn: psycopg.Connection[Any]) -> None:
        with conn.cursor() as cur:
            cur.execute(sql.SQL("CREATE SCHEMA IF NOT EXISTS {}").format(sql.Identifier(self.schema)))
            cur.execute(
                sql.SQL(
                    """
                    CREATE TABLE IF NOT EXISTS {}.import_batches (
                        id BIGSERIAL PRIMARY KEY,
                        dataset_name VARCHAR(80) NOT NULL,
                        source_file TEXT NOT NULL,
                        file_hash VARCHAR(64) NOT NULL,
                        reference_date DATE,
                        total_rows INTEGER NOT NULL,
                        imported_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                    """
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    """
                    CREATE TABLE IF NOT EXISTS {}.critica_rn_snapshot (
                        batch_id BIGINT NOT NULL REFERENCES {}.import_batches(id) ON DELETE CASCADE,
                        row_number BIGINT NOT NULL,
                        filial VARCHAR(16) NOT NULL DEFAULT '',
                        pedido VARCHAR(32) NOT NULL DEFAULT '',
                        data_pedido DATE,
                        operacao VARCHAR(32) NOT NULL DEFAULT '',
                        cod_pdv VARCHAR(32) NOT NULL DEFAULT '',
                        nome_pdv TEXT NOT NULL DEFAULT '',
                        setor VARCHAR(32) NOT NULL DEFAULT '',
                        filial_setor_key TEXT NOT NULL DEFAULT '',
                        filial_gv_key TEXT NOT NULL DEFAULT '',
                        filial_dc_key TEXT NOT NULL DEFAULT '',
                        status_pedido TEXT NOT NULL DEFAULT '',
                        total_pedido NUMERIC(18, 4) NOT NULL DEFAULT 0,
                        total_cliente NUMERIC(18, 4) NOT NULL DEFAULT 0,
                        critica_text TEXT NOT NULL DEFAULT '',
                        produto_codigo VARCHAR(32) NOT NULL DEFAULT '',
                        produto_key VARCHAR(32) NOT NULL DEFAULT '',
                        quantidade NUMERIC(18, 4) NOT NULL DEFAULT 0,
                        unid_venda VARCHAR(32) NOT NULL DEFAULT '',
                        preco_unitario NUMERIC(18, 4) NOT NULL DEFAULT 0,
                        preco_sem_adf NUMERIC(18, 4) NOT NULL DEFAULT 0,
                        minimo_politica NUMERIC(18, 4) NOT NULL DEFAULT 0,
                        tipo_movimento VARCHAR(32) NOT NULL DEFAULT '',
                        codigo_gv VARCHAR(32) NOT NULL DEFAULT '',
                        codigo_pgv VARCHAR(32) NOT NULL DEFAULT '',
                        payload JSONB NOT NULL,
                        imported_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        PRIMARY KEY (batch_id, row_number)
                    )
                    """
                ).format(sql.Identifier(self.schema), sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS critica_rn_snapshot_batch_filial_setor_idx ON {}.critica_rn_snapshot (batch_id, filial_setor_key)"
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS critica_rn_snapshot_batch_pedido_idx ON {}.critica_rn_snapshot (batch_id, filial, pedido)"
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS critica_rn_snapshot_batch_data_idx ON {}.critica_rn_snapshot (batch_id, data_pedido)"
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS critica_rn_snapshot_batch_cliente_idx ON {}.critica_rn_snapshot (batch_id, filial, cod_pdv, data_pedido, pedido)"
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS critica_rn_snapshot_batch_gv_idx ON {}.critica_rn_snapshot (batch_id, filial_gv_key)"
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS critica_rn_snapshot_batch_dc_idx ON {}.critica_rn_snapshot (batch_id, filial_dc_key)"
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS critica_rn_snapshot_batch_produto_idx ON {}.critica_rn_snapshot (batch_id, produto_codigo)"
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS import_batches_critica_rn_dataset_idx ON {}.import_batches (dataset_name, imported_at DESC)"
                ).format(sql.Identifier(self.schema))
            )
            ensure_dataset_state_table(conn, self.schema)

    def _insert_batch(
        self,
        conn: psycopg.Connection[Any],
        source_file: str,
        reference_date: date,
        file_hash: str,
        total_rows: int,
    ) -> int:
        query = sql.SQL(
            """
            INSERT INTO {}.import_batches (dataset_name, source_file, file_hash, reference_date, total_rows)
            VALUES ('critica_rn', %s, %s, %s, %s)
            RETURNING id
            """
        ).format(sql.Identifier(self.schema))
        with conn.cursor(row_factory=dict_row) as cur:
            cur.execute(query, (source_file, file_hash, reference_date, total_rows))
            row = cur.fetchone()
        return int(row["id"])

    def _insert_snapshot_rows(
        self,
        conn: psycopg.Connection[Any],
        rows: list[CriticaRnRow],
        batch_id: int,
    ) -> None:
        query = sql.SQL(
            """
            INSERT INTO {}.critica_rn_snapshot (
                batch_id,
                row_number,
                filial,
                pedido,
                data_pedido,
                operacao,
                cod_pdv,
                nome_pdv,
                setor,
                filial_setor_key,
                status_pedido,
                total_pedido,
                total_cliente,
                critica_text,
                produto_codigo,
                produto_key,
                quantidade,
                unid_venda,
                preco_unitario,
                preco_sem_adf,
                minimo_politica,
                tipo_movimento,
                codigo_gv,
                codigo_pgv,
                payload
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
        ).format(sql.Identifier(self.schema))
        payload_batch: list[tuple[Any, ...]] = []
        with conn.cursor() as cur:
            for row_number, row in enumerate(rows, start=1):
                payload_batch.append(
                    (
                        batch_id,
                        row_number,
                        row.filial,
                        row.pedido,
                        row.data_pedido,
                        row.operacao,
                        row.cod_pdv,
                        row.nome_pdv,
                        row.setor,
                        row.filial_setor_key,
                        row.status_pedido,
                        row.total_pedido,
                        row.total_cliente,
                        row.critica_text,
                        row.produto_codigo,
                        row.produto_key,
                        row.quantidade,
                        row.unid_venda,
                        row.preco_unitario,
                        row.preco_sem_adf,
                        row.minimo_politica,
                        row.tipo_movimento,
                        row.codigo_gv,
                        row.codigo_pgv,
                        Jsonb(row.payload),
                    )
                )
                if len(payload_batch) >= 500:
                    cur.executemany(query, payload_batch)
                    payload_batch.clear()
            if payload_batch:
                cur.executemany(query, payload_batch)

    def _hydrate_scope_columns(self, conn: psycopg.Connection[Any], batch_id: int) -> None:
        if not _relation_exists(conn, self.schema, "dsetores_latest"):
            return
        query = sql.SQL(
            """
            UPDATE {schema}.critica_rn_snapshot AS c
            SET
                filial_gv_key = COALESCE(ds.filial_gv_key, ''),
                filial_dc_key = COALESCE(ds.filial_dc_key, '')
            FROM {schema}.dsetores_latest AS ds
            WHERE c.batch_id = %s
              AND ds.filial_setor_key = c.filial_setor_key
            """
        ).format(schema=sql.Identifier(self.schema))
        with conn.cursor() as cur:
            cur.execute(query, (batch_id,))

    def _create_latest_view(self, conn: psycopg.Connection[Any]) -> None:
        active_batch_id = resolve_effective_import_batch_id(conn, self.schema, "critica_rn", activate_if_missing=True)
        where_clause = sql.SQL("c.batch_id = {}").format(sql.Literal(active_batch_id)) if active_batch_id is not None else sql.SQL("FALSE")
        price_ref_cte = _price_reference_cte(self.schema, conn)
        query = sql.SQL(
            """
            CREATE OR REPLACE VIEW {}.critica_rn_latest AS
            WITH
            {},
            base AS (
                SELECT
                    c.*,
                    COUNT(*) OVER (PARTITION BY c.batch_id, c.filial, c.pedido) AS pedido_linhas,
                    COUNT(*) OVER (PARTITION BY c.batch_id, c.filial, c.pedido, c.produto_codigo) AS pedido_produto_linhas
                FROM {}.critica_rn_snapshot c
                WHERE {}
            )
            SELECT
                base.batch_id,
                base.row_number,
                base.filial,
                base.pedido,
                CONCAT(base.filial, '_', base.pedido) AS filial_pedido_key,
                base.data_pedido,
                base.operacao,
                base.cod_pdv,
                base.nome_pdv,
                base.setor,
                base.filial_setor_key,
                base.filial_gv_key,
                base.filial_dc_key,
                base.status_pedido,
                base.total_pedido,
                base.total_cliente,
                base.critica_text,
                BTRIM(COALESCE(base.critica_text, '')) <> '' AS possui_critica,
                base.produto_codigo,
                base.produto_key,
                CONCAT(base.filial, '_', base.pedido, '_', base.produto_codigo) AS filial_pedido_produto_key,
                base.quantidade,
                base.unid_venda,
                base.preco_unitario,
                base.preco_sem_adf,
                base.minimo_politica,
                base.tipo_movimento,
                base.codigo_gv,
                base.codigo_pgv,
                base.pedido_linhas,
                base.pedido_produto_linhas,
                base.pedido_produto_linhas > 1 AS pedido_produto_duplicado,
                dp.produto_dprecos,
                dp.dprecos_match_count,
                dp.ttc_min,
                dp.ttc_max,
                dp.caixa_min,
                dp.caixa_max,
                CASE
                    WHEN BTRIM(COALESCE(base.produto_codigo, '')) = '' THEN FALSE
                    ELSE dp.codigo IS NOT NULL
                END AS produto_encontrado_dprecos,
                CASE
                    WHEN BTRIM(COALESCE(base.produto_codigo, '')) = '' THEN 'sem_produto'
                    WHEN dp.codigo IS NULL THEN 'produto_sem_dprecos'
                    WHEN LOWER(BTRIM(COALESCE(base.unid_venda, ''))) IN ('un', 'und', 'unid', 'unidade')
                         AND dp.ttc_min IS NOT NULL
                         AND (base.preco_unitario < dp.ttc_min - 0.01 OR base.preco_unitario > dp.ttc_max + 0.01)
                        THEN 'preco_unidade_fora_referencia'
                    WHEN LOWER(BTRIM(COALESCE(base.unid_venda, ''))) NOT IN ('un', 'und', 'unid', 'unidade')
                         AND dp.caixa_min IS NOT NULL
                         AND (base.preco_unitario < dp.caixa_min - 0.01 OR base.preco_unitario > dp.caixa_max + 0.01)
                        THEN 'preco_caixa_fora_referencia'
                    ELSE 'ok'
                END AS preco_status,
                base.payload,
                base.imported_at,
                b.reference_date,
                b.source_file,
                b.file_hash,
                b.imported_at AS batch_imported_at
            FROM base
            JOIN {}.import_batches b ON b.id = base.batch_id
            LEFT JOIN dprecos_ref dp ON dp.codigo = base.produto_codigo
            """
        ).format(
            sql.Identifier(self.schema),
            price_ref_cte,
            sql.Identifier(self.schema),
            where_clause,
            sql.Identifier(self.schema),
        )
        with conn.cursor() as cur:
            cur.execute(query)

    def _connect(self) -> psycopg.Connection[Any]:
        return psycopg.connect(self.database_url, connect_timeout=self.connect_timeout_seconds)


def _load_critica_rn_rows(path: Path) -> list[CriticaRnRow]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _load_critica_rn_rows_from_workbook(path)
    return _load_critica_rn_rows_from_csv(path)


def _load_critica_rn_rows_from_csv(path: Path) -> list[CriticaRnRow]:
    text = _read_text_with_fallback(path)
    delimiter = _detect_delimiter(text)
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    if reader.fieldnames is None:
        raise ValueError("Arquivo sem cabecalho.")
    headers = [str(header or "").strip() for header in reader.fieldnames]
    header_map = _build_header_map(headers)
    raw_rows = [
        {str(key or "").strip(): str(value or "").strip() for key, value in raw_row.items() if key is not None}
        for raw_row in reader
    ]
    return _build_critica_rn_rows_from_mapping_rows(raw_rows, headers=headers, header_map=header_map, row_number_offset=2)


def _load_critica_rn_rows_from_workbook(path: Path) -> list[CriticaRnRow]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_values is None:
            raise ValueError("Planilha sem cabecalho.")
        headers = [str(value or "").strip() for value in header_values]
        header_map = _build_header_map(headers)
        rows_as_mapping: list[dict[str, Any]] = []
        for row_values in worksheet.iter_rows(min_row=2, values_only=True):
            row_map: dict[str, Any] = {}
            for index, header in enumerate(headers):
                key = header or f"col_{index + 1:03d}"
                value = row_values[index] if index < len(row_values) else None
                row_map[key] = value
            rows_as_mapping.append(row_map)
        return _build_critica_rn_rows_from_mapping_rows(rows_as_mapping, headers=headers, header_map=header_map, row_number_offset=2)
    finally:
        workbook.close()


def _build_header_map(headers: list[str]) -> dict[str, str]:
    header_map: dict[str, str] = {}
    for actual in headers:
        normalized = _normalize_header(actual)
        canonical = HEADER_ALIASES.get(normalized)
        if canonical and canonical not in header_map:
            header_map[canonical] = actual

    missing_headers = sorted(EXPECTED_HEADERS - set(header_map))
    if missing_headers:
        raise ValueError(f"Arquivo invalido. Colunas obrigatorias ausentes: {', '.join(missing_headers)}")
    return header_map


def _build_critica_rn_rows_from_mapping_rows(
    raw_rows: list[dict[str, Any]],
    *,
    headers: list[str],
    header_map: dict[str, str],
    row_number_offset: int,
) -> list[CriticaRnRow]:
    rows: list[CriticaRnRow] = []
    for index, row in enumerate(raw_rows, start=row_number_offset):
        if not any(str(value or "").strip() for value in row.values()):
            continue
        filial = normalize_numeric_code(row.get(header_map["UNB"], ""))
        pedido = normalize_numeric_code(row.get(header_map["Pedido"], ""))
        setor = normalize_numeric_code(row.get(header_map["Setor"], ""))
        produto_codigo = normalize_numeric_code(row.get(header_map["Produto"], ""))
        cod_pdv = normalize_numeric_code(row.get(header_map["Cod. PDV"], ""))
        if not filial or not pedido or not setor or not produto_codigo:
            continue
        criticas = [
            _clean_text(row.get(header_map[key], ""))
            for key in ("Critica 1", "Critica 2", "Critica 3", "Critica 4", "Critica 5", "Critica 6", "Critica Rejeicao")
            if key in header_map
        ]
        rows.append(
            CriticaRnRow(
                filial=filial,
                pedido=pedido,
                data_pedido=_parse_date(row.get(header_map["Data Pedido"], "")),
                operacao=normalize_numeric_code(row.get(header_map.get("Operacao", ""), "")),
                cod_pdv=cod_pdv,
                nome_pdv=_clean_text(row.get(header_map["Nome PDV"], "")),
                setor=setor,
                filial_setor_key=f"{filial}_{setor}",
                status_pedido=_clean_text(row.get(header_map.get("Status Pedido", ""), "")),
                total_pedido=_parse_decimal_value(row.get(header_map.get("Total Pedido", ""), "")),
                total_cliente=_parse_decimal_value(row.get(header_map.get("Total Cliente", ""), "")),
                critica_text=" | ".join(item for item in criticas if item),
                produto_codigo=produto_codigo,
                produto_key=produto_codigo,
                quantidade=_parse_decimal_value(row.get(header_map["Quantidade"], "")),
                unid_venda=_clean_text(row.get(header_map["Unid. Venda"], "")),
                preco_unitario=_parse_decimal_value(row.get(header_map["Preco Unitario"], "")),
                preco_sem_adf=_parse_decimal_value(row.get(header_map.get("Preco S/ ADF", ""), "")),
                minimo_politica=_parse_decimal_value(row.get(header_map.get("Minimo Politica", ""), "")),
                tipo_movimento=normalize_numeric_code(row.get(header_map.get("Tipo Movimento", ""), "")),
                codigo_gv=normalize_numeric_code(row.get(header_map.get("Codigo GV", ""), "")),
                codigo_pgv=normalize_numeric_code(row.get(header_map.get("Codigo PGV", ""), "")),
                source_row_number=index,
                payload=_row_to_payload(headers, row),
            )
        )
    return rows


def _summary_stats(rows: list[CriticaRnRow]) -> dict[str, int]:
    pedido_keys = {(row.filial, row.pedido) for row in rows}
    unb_setores = {row.filial_setor_key for row in rows if row.filial_setor_key}
    produtos = {row.produto_codigo for row in rows if row.produto_codigo}
    pedido_produto_counts: dict[tuple[str, str, str], int] = {}
    for row in rows:
        key = (row.filial, row.pedido, row.produto_codigo)
        pedido_produto_counts[key] = pedido_produto_counts.get(key, 0) + 1
    return {
        "unique_pedidos": len(pedido_keys),
        "unique_unb_setores": len(unb_setores),
        "unique_produtos": len(produtos),
        "duplicate_pedido_produto_keys": sum(1 for total in pedido_produto_counts.values() if total > 1),
        "rows_with_critica": sum(1 for row in rows if row.critica_text),
    }


def _row_to_payload(headers: list[str], row: dict[str, Any]) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for index, header in enumerate(headers, start=1):
        key = header or f"col_{index:03d}"
        payload[key] = _cell_to_json(row.get(key, ""))
    return payload


def _cell_to_json(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, float):
        return f"{value:.10f}".rstrip("0").rstrip(".") or "0"
    if value is None:
        return ""
    return str(value).strip()


def _read_text_with_fallback(path: Path) -> str:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("Nao consegui ler o arquivo com um encoding suportado.")


def _detect_delimiter(text: str) -> str:
    sample = "\n".join(text.splitlines()[:10])
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;")
        return str(dialect.delimiter or ",")
    except Exception:
        return ","


def _normalize_header(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_only = "".join(char for char in normalized if not unicodedata.combining(char))
    return "".join(char for char in ascii_only.lower() if char.isalnum())


def _normalize_schema(value: str) -> str:
    cleaned = "".join(char for char in str(value or "").strip() if char.isalnum() or char == "_")
    return cleaned or "reports"


def _clean_text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def _parse_decimal_value(value: Any) -> Decimal:
    cleaned = str(value or "").strip()
    if not cleaned:
        return Decimal("0")
    normalized = cleaned.replace("R$", "").replace(" ", "")
    if "," in normalized and "." in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")
    elif "," in normalized:
        normalized = normalized.replace(",", ".")
    try:
        return Decimal(normalized)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    cleaned = str(value or "").strip()
    if not cleaned:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(cleaned, fmt).date()
        except ValueError:
            continue
    return None


def _relation_exists(conn: psycopg.Connection[Any], schema: str, relation: str) -> bool:
    with conn.cursor(row_factory=dict_row) as cur:
        cur.execute("SELECT to_regclass(%s) AS relation_name", (f"{schema}.{relation}",))
        row = cur.fetchone()
    return bool(row and row["relation_name"])


def _price_reference_cte(schema: str, conn: psycopg.Connection[Any]) -> sql.Composed:
    if not _relation_exists(conn, schema, "dprecos_latest"):
        return sql.SQL(
            """
            dprecos_ref AS (
                SELECT
                    NULL::text AS codigo,
                    NULL::text AS produto_dprecos,
                    0::bigint AS dprecos_match_count,
                    NULL::numeric AS ttc_min,
                    NULL::numeric AS ttc_max,
                    NULL::numeric AS caixa_min,
                    NULL::numeric AS caixa_max
                WHERE FALSE
            )
            """
        )
    return sql.SQL(
        """
        dprecos_ref AS (
            SELECT
                p.codigo,
                MIN(p.produto) AS produto_dprecos,
                COUNT(DISTINCT p.row_number) AS dprecos_match_count,
                MIN(NULLIF(p.ttc, 0)) AS ttc_min,
                MAX(NULLIF(p.ttc, 0)) AS ttc_max,
                MIN(NULLIF(v.caixa_preco, 0)) AS caixa_min,
                MAX(NULLIF(v.caixa_preco, 0)) AS caixa_max
            FROM {}.dprecos_latest p
            LEFT JOIN LATERAL (VALUES (p.asr), (p.sub), (p.frio)) AS v(caixa_preco) ON TRUE
            GROUP BY p.codigo
        )
        """
    ).format(sql.Identifier(schema))


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fp:
        for chunk in iter(lambda: fp.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
