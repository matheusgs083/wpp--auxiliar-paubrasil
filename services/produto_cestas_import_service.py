from __future__ import annotations

import csv
import hashlib
import io
import unicodedata
import warnings
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import psycopg
from psycopg import sql
from psycopg.rows import dict_row
from psycopg.types.json import Jsonb

from bot_api.commercial_scope import normalize_numeric_code
from bot_api.services.import_publication import (
    activate_import_batch,
    ensure_dataset_state_table,
    prune_import_batches,
    resolve_effective_import_batch_id,
)


EXPECTED_EXTENSION_SET = {".xlsx", ".xlsm", ".csv"}
REQUIRED_HEADERS = {"codigo", "nome_cesta", "nome_produto"}
HEADER_ALIASES = {
    "codigoabreviadoproduto": "codigo",
    "codigo": "codigo",
    "codigoproduto": "codigo",
    "codproduto": "codigo",
    "nomecesta": "nome_cesta",
    "cesta": "nome_cesta",
    "categoria": "nome_cesta",
    "nomeproduto": "nome_produto",
    "produto": "nome_produto",
    "nomefornecmktplace": "nome_fornecedor_marketplace",
    "nomefornecedormktplace": "nome_fornecedor_marketplace",
    "fornecedormarketplace": "nome_fornecedor_marketplace",
}


@dataclass(frozen=True)
class ProdutoCestasValidationResult:
    dataset_name: str
    source_path: str
    ok: bool
    total_rows: int
    unique_codigos: int
    unique_cestas: int
    unique_categorias: int
    error_count: int
    warning_count: int
    sample_errors: list[str]
    sample_warnings: list[str]

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)

    def ensure_valid(self) -> None:
        if self.error_count:
            raise ValueError("; ".join(self.sample_errors) if self.sample_errors else "Erros na validacao da cesta de produtos.")


@dataclass(frozen=True)
class ProdutoCestasImportSummary:
    source_path: str
    rows: int
    unique_codigos: int
    unique_cestas: int
    unique_categorias: int

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class ProdutoCestaRow:
    codigo: str
    nome_cesta: str
    nome_produto: str
    nome_fornecedor_marketplace: str
    categoria_tipo: str
    categoria_nome: str
    source_row_number: int
    payload: dict[str, str]


class ProdutoCestasImportService:
    def __init__(self, database_url: str, schema: str, connect_timeout_seconds: float = 3.0) -> None:
        self.database_url = database_url.strip()
        self.schema = _normalize_schema(schema)
        self.connect_timeout_seconds = max(float(connect_timeout_seconds), 1.0)

    def validate_source(self, source_path: Path) -> ProdutoCestasValidationResult:
        path = source_path.expanduser().resolve()
        errors: list[str] = []
        warnings_list: list[str] = []
        rows: list[ProdutoCestaRow] = []

        if not path.exists():
            errors.append(f"Arquivo nao encontrado: {path}")
        elif path.suffix.lower() not in EXPECTED_EXTENSION_SET:
            errors.append("Formato invalido. Use um arquivo .xlsx, .xlsm ou .csv.")
        else:
            try:
                rows = _load_produto_cesta_rows(path)
            except Exception as exc:
                errors.append(str(exc))

        if not rows and not errors:
            errors.append("Nao encontrei linhas validas na cesta de produtos.")

        return ProdutoCestasValidationResult(
            dataset_name="produto_cestas",
            source_path=str(path),
            ok=not errors,
            total_rows=len(rows),
            unique_codigos=len({row.codigo for row in rows}),
            unique_cestas=len({row.nome_cesta for row in rows if row.nome_cesta}),
            unique_categorias=len({row.categoria_nome for row in rows if row.categoria_nome}),
            error_count=len(errors),
            warning_count=len(warnings_list),
            sample_errors=errors[:10],
            sample_warnings=warnings_list[:10],
        )

    def summarize_source(self, source_path: Path) -> ProdutoCestasImportSummary:
        validation = self.validate_source(source_path)
        validation.ensure_valid()
        rows = _load_produto_cesta_rows(source_path.expanduser().resolve())
        return ProdutoCestasImportSummary(
            source_path=str(source_path),
            rows=len(rows),
            unique_codigos=len({row.codigo for row in rows}),
            unique_cestas=len({row.nome_cesta for row in rows if row.nome_cesta}),
            unique_categorias=len({row.categoria_nome for row in rows if row.categoria_nome}),
        )

    def import_source(self, source_path: Path, reference_date: date | None = None) -> dict[str, Any]:
        if not self.database_url:
            raise RuntimeError("REPORTS_DATABASE_URL nao configurada.")

        path = source_path.expanduser().resolve()
        validation = self.validate_source(path)
        validation.ensure_valid()
        summary = self.summarize_source(path)
        rows = _load_produto_cesta_rows(path)
        batch_date = reference_date or datetime.fromtimestamp(path.stat().st_mtime).date()
        source_hash = _sha256(path)

        with self._connect() as conn:
            self._ensure_schema(conn)
            batch_id = self._insert_batch(conn, str(path), batch_date, source_hash, len(rows))
            self._insert_snapshot_rows(conn, rows, batch_id)
            activate_import_batch(conn, self.schema, "produto_cestas", batch_id)
            self._create_latest_views(conn)
            prune_import_batches(conn, self.schema, "produto_cestas", keep_last=3)
            conn.commit()

        result = summary.to_dict()
        result.update(
            {
                "batch_id": batch_id,
                "reference_date": batch_date.isoformat(),
                "source_hash": source_hash,
                "schema": self.schema,
                "replaced_previous_batches": False,
                "published_as_active_batch": True,
            }
        )
        return result

    def refresh_latest_view(self) -> dict[str, Any]:
        if not self.database_url:
            raise RuntimeError("REPORTS_DATABASE_URL nao configurada.")

        with self._connect() as conn:
            self._ensure_schema(conn)
            active_batch_id = resolve_effective_import_batch_id(conn, self.schema, "produto_cestas", activate_if_missing=True)
            self._create_latest_views(conn)
            conn.commit()

        return {
            "ok": True,
            "schema": self.schema,
            "view": f"{self.schema}.produto_cestas_latest",
            "category_view": f"{self.schema}.produto_categorias_latest",
            "active_batch_id": active_batch_id,
        }

    def _ensure_schema(self, conn: psycopg.Connection[Any]) -> None:
        with conn.cursor() as cur:
            cur.execute(sql.SQL("CREATE SCHEMA IF NOT EXISTS {}").format(sql.Identifier(self.schema)))
            cur.execute(
                sql.SQL(
                    """
                    CREATE TABLE IF NOT EXISTS {}.import_batches (
                        id BIGSERIAL PRIMARY KEY,
                        dataset_name VARCHAR(80) NOT NULL,
                        source_file TEXT NOT NULL,
                        file_hash VARCHAR(64) NOT NULL,
                        reference_date DATE,
                        total_rows INTEGER NOT NULL,
                        imported_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                    """
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    """
                    CREATE TABLE IF NOT EXISTS {}.produto_cestas_snapshot (
                        batch_id BIGINT NOT NULL REFERENCES {}.import_batches(id) ON DELETE CASCADE,
                        row_number BIGINT NOT NULL,
                        codigo VARCHAR(32) NOT NULL,
                        nome_cesta TEXT NOT NULL,
                        nome_produto TEXT NOT NULL,
                        nome_fornecedor_marketplace TEXT NOT NULL DEFAULT '',
                        categoria_tipo VARCHAR(64) NOT NULL DEFAULT '',
                        categoria_nome TEXT NOT NULL DEFAULT '',
                        payload JSONB NOT NULL,
                        imported_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        PRIMARY KEY (batch_id, row_number)
                    )
                    """
                ).format(sql.Identifier(self.schema), sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS produto_cestas_snapshot_batch_codigo_idx ON {}.produto_cestas_snapshot (batch_id, codigo)"
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS produto_cestas_snapshot_batch_tipo_idx ON {}.produto_cestas_snapshot (batch_id, categoria_tipo)"
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS import_batches_produto_cestas_dataset_idx ON {}.import_batches (dataset_name, imported_at DESC)"
                ).format(sql.Identifier(self.schema))
            )
            ensure_dataset_state_table(conn, self.schema)

    def _insert_batch(
        self,
        conn: psycopg.Connection[Any],
        source_file: str,
        reference_date: date,
        file_hash: str,
        total_rows: int,
    ) -> int:
        query = sql.SQL(
            """
            INSERT INTO {}.import_batches (dataset_name, source_file, file_hash, reference_date, total_rows)
            VALUES ('produto_cestas', %s, %s, %s, %s)
            RETURNING id
            """
        ).format(sql.Identifier(self.schema))
        with conn.cursor(row_factory=dict_row) as cur:
            cur.execute(query, (source_file, file_hash, reference_date, total_rows))
            row = cur.fetchone()
        return int(row["id"])

    def _insert_snapshot_rows(
        self,
        conn: psycopg.Connection[Any],
        rows: list[ProdutoCestaRow],
        batch_id: int,
    ) -> None:
        query = sql.SQL(
            """
            INSERT INTO {}.produto_cestas_snapshot (
                batch_id,
                row_number,
                codigo,
                nome_cesta,
                nome_produto,
                nome_fornecedor_marketplace,
                categoria_tipo,
                categoria_nome,
                payload
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
        ).format(sql.Identifier(self.schema))
        params = [
            (
                batch_id,
                index,
                row.codigo,
                row.nome_cesta,
                row.nome_produto,
                row.nome_fornecedor_marketplace,
                row.categoria_tipo,
                row.categoria_nome,
                Jsonb(row.payload),
            )
            for index, row in enumerate(rows, start=1)
        ]
        with conn.cursor() as cur:
            cur.executemany(query, params)

    def _create_latest_views(self, conn: psycopg.Connection[Any]) -> None:
        active_batch_id = resolve_effective_import_batch_id(conn, self.schema, "produto_cestas", activate_if_missing=True)
        where_clause = sql.SQL("s.batch_id = {}").format(sql.Literal(active_batch_id)) if active_batch_id is not None else sql.SQL("FALSE")
        latest_query = sql.SQL(
            """
            CREATE OR REPLACE VIEW {}.produto_cestas_latest AS
            SELECT
                s.batch_id,
                s.row_number,
                s.codigo,
                s.nome_cesta,
                s.nome_produto,
                s.nome_fornecedor_marketplace,
                s.categoria_tipo,
                s.categoria_nome,
                s.payload,
                s.imported_at,
                b.reference_date,
                b.source_file,
                b.file_hash,
                b.imported_at AS batch_imported_at
            FROM {}.produto_cestas_snapshot s
            JOIN {}.import_batches b ON b.id = s.batch_id
            WHERE {}
            """
        ).format(
            sql.Identifier(self.schema),
            sql.Identifier(self.schema),
            sql.Identifier(self.schema),
            where_clause,
        )
        category_query = sql.SQL(
            """
            CREATE OR REPLACE VIEW {}.produto_categorias_latest AS
            SELECT
                c.batch_id,
                c.codigo,
                MAX(c.nome_produto) AS nome_produto,
                MAX(c.categoria_nome) FILTER (WHERE c.categoria_tipo = 'categoria') AS categoria,
                MAX(c.categoria_nome) FILTER (WHERE c.categoria_tipo = 'categoria_agrupada') AS categoria_agrupada,
                STRING_AGG(DISTINCT c.categoria_nome, '; ' ORDER BY c.categoria_nome) FILTER (WHERE c.categoria_tipo = 'familia') AS familias,
                STRING_AGG(DISTINCT c.nome_cesta, '; ' ORDER BY c.nome_cesta) AS cestas,
                MAX(c.reference_date) AS reference_date,
                MAX(c.source_file) AS source_file,
                MAX(c.file_hash) AS file_hash,
                MAX(c.batch_imported_at) AS batch_imported_at
            FROM {}.produto_cestas_latest c
            GROUP BY c.batch_id, c.codigo
            """
        ).format(sql.Identifier(self.schema), sql.Identifier(self.schema))
        with conn.cursor() as cur:
            cur.execute(latest_query)
            cur.execute(category_query)

    def _connect(self) -> psycopg.Connection[Any]:
        return psycopg.connect(self.database_url, connect_timeout=self.connect_timeout_seconds)


def _load_produto_cesta_rows(path: Path) -> list[ProdutoCestaRow]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _load_produto_cesta_rows_from_workbook(path)
    return _load_produto_cesta_rows_from_csv(path)


def _load_produto_cesta_rows_from_csv(path: Path) -> list[ProdutoCestaRow]:
    text = _read_text_with_fallback(path)
    delimiter = _detect_delimiter(text)
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    if reader.fieldnames is None:
        raise ValueError("Arquivo sem cabecalho.")
    header_row = [str(header or "").strip() for header in reader.fieldnames]
    header_map = _build_header_map(header_row)
    raw_rows = [
        {str(key or "").strip(): str(value or "").strip() for key, value in raw_row.items() if key is not None}
        for raw_row in reader
    ]
    return _build_produto_cesta_rows_from_mapping_rows(raw_rows, headers=header_row, header_map=header_map, row_number_offset=2)


def _load_produto_cesta_rows_from_workbook(path: Path) -> list[ProdutoCestaRow]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        iterator = worksheet.iter_rows(values_only=True)
        header_values = next(iterator, None)
        if header_values is None:
            raise ValueError("Planilha sem cabecalho.")

        header_row = [str(value or "").strip() for value in header_values]
        header_map = _build_header_map(header_row)
        rows_as_mapping: list[dict[str, str]] = []
        for row_values in iterator:
            row_map: dict[str, str] = {}
            for index, header in enumerate(header_row):
                if not header:
                    continue
                value = row_values[index] if index < len(row_values) else None
                row_map[header] = _cell_to_text(value)
            rows_as_mapping.append(row_map)
        return _build_produto_cesta_rows_from_mapping_rows(rows_as_mapping, headers=header_row, header_map=header_map, row_number_offset=2)
    finally:
        workbook.close()


def _build_produto_cesta_rows_from_mapping_rows(
    raw_rows: list[dict[str, str]],
    *,
    headers: list[str],
    header_map: dict[str, str],
    row_number_offset: int,
) -> list[ProdutoCestaRow]:
    rows: list[ProdutoCestaRow] = []
    for index, row in enumerate(raw_rows, start=row_number_offset):
        if not any(str(value or "").strip() for value in row.values()):
            continue
        codigo = normalize_numeric_code(row.get(header_map["codigo"], ""))
        nome_cesta = _clean_text(row.get(header_map["nome_cesta"], ""))
        nome_produto = _clean_text(row.get(header_map["nome_produto"], ""))
        if not codigo or not nome_cesta or not nome_produto:
            continue
        categoria_tipo, categoria_nome = _parse_categoria_from_cesta(nome_cesta)
        rows.append(
            ProdutoCestaRow(
                codigo=codigo,
                nome_cesta=nome_cesta,
                nome_produto=nome_produto,
                nome_fornecedor_marketplace=_clean_text(row.get(header_map.get("nome_fornecedor_marketplace", ""), "")),
                categoria_tipo=categoria_tipo,
                categoria_nome=categoria_nome,
                source_row_number=index,
                payload={header: _clean_text(row.get(header, "")) for header in headers if header},
            )
        )
    rows.sort(key=lambda item: (_sort_numeric(item.codigo), item.categoria_tipo, item.nome_cesta, item.source_row_number))
    return rows


def _build_header_map(headers: list[str]) -> dict[str, str]:
    header_map: dict[str, str] = {}
    for actual in headers:
        normalized = _normalize_header(actual)
        canonical = HEADER_ALIASES.get(normalized)
        if canonical and canonical not in header_map:
            header_map[canonical] = actual
    missing_headers = sorted(REQUIRED_HEADERS - set(header_map))
    if missing_headers:
        raise ValueError(f"Arquivo invalido. Colunas obrigatorias ausentes: {', '.join(missing_headers)}")
    return header_map


def _parse_categoria_from_cesta(nome_cesta: str) -> tuple[str, str]:
    text = _clean_text(nome_cesta)
    if " - " not in text:
        return "cesta", text
    left, right = (part.strip() for part in text.split(" - ", 1))
    left_norm = _normalize_header(left)
    right_norm = _normalize_header(right)
    if left_norm == "categoria":
        return "categoria", right
    if left_norm == "categoriaagrupado":
        return "categoria_agrupada", right
    if right_norm == "familia":
        return "familia", left
    return "cesta", right or text


def _read_text_with_fallback(path: Path) -> str:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("Nao consegui ler o arquivo com um encoding suportado.")


def _detect_delimiter(text: str) -> str:
    sample = "\n".join(text.splitlines()[:10])
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;")
        return str(dialect.delimiter or ";")
    except Exception:
        return ";"


def _normalize_header(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_only = "".join(char for char in normalized if not unicodedata.combining(char))
    return "".join(char for char in ascii_only.lower() if char.isalnum())


def _normalize_schema(value: str) -> str:
    cleaned = "".join(char for char in str(value or "").strip() if char.isalnum() or char == "_")
    return cleaned or "reports"


def _clean_text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        text = f"{value:.10f}".rstrip("0").rstrip(".")
        return text or "0"
    return str(value).strip()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fp:
        for chunk in iter(lambda: fp.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _sort_numeric(value: str) -> tuple[int, str]:
    cleaned = normalize_numeric_code(value)
    if cleaned:
        try:
            return (0, f"{int(cleaned):09d}")
        except ValueError:
            return (1, cleaned)
    return (2, str(value or "").strip())
