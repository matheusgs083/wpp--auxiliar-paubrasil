from __future__ import annotations

import csv
import hashlib
import io
import unicodedata
import warnings
from dataclasses import asdict, dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import psycopg
from psycopg import sql
from psycopg.rows import dict_row
from psycopg.types.json import Jsonb

from bot_api.commercial_scope import normalize_numeric_code
from bot_api.services.import_publication import (
    activate_import_batch,
    ensure_dataset_state_table,
    prune_import_batches,
    resolve_effective_import_batch_id,
)


EXPECTED_EXTENSION_SET = {".xlsx", ".xlsm", ".xls", ".csv"}
EXPECTED_HEADERS = {
    "KPI",
    "Filial",
    "NB",
    "% Pag Atraso",
    "Prazo Atual",
    "Cond Pag Atual",
    "Limite Total",
    "Faturamento com PDV",
}
OPTIONAL_HEADERS = {
    "Pedidos",
}
HEADER_ALIASES = {
    "kpi": "KPI",
    "filial": "Filial",
    "nb": "NB",
    "pagatraso": "% Pag Atraso",
    "prazoatual": "Prazo Atual",
    "condpagatual": "Cond Pag Atual",
    "limitetotal": "Limite Total",
    "faturamentocompdv": "Faturamento com PDV",
    "pedidos": "Pedidos",
}


@dataclass(frozen=True)
class PrazoLimiteValidationResult:
    dataset_name: str
    source_path: str
    ok: bool
    total_rows: int
    unique_clientes: int
    error_count: int
    warning_count: int
    sample_errors: list[str]
    sample_warnings: list[str]

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)

    def ensure_valid(self) -> None:
        if self.error_count:
            raise ValueError("; ".join(self.sample_errors) if self.sample_errors else "Erros na validacao da base de prazo e limite.")


@dataclass(frozen=True)
class PrazoLimiteImportSummary:
    source_path: str
    rows: int
    unique_clientes: int
    unique_filiais: int
    unique_periodos: int
    limite_total_somado: str
    faturamento_total_somado: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class PrazoLimiteRow:
    filial: str
    cod_pdv: str
    chave: str
    kpi: str
    percentual_pag_atraso: str
    prazo_atual: str
    cond_pag_atual: str
    limite_total: Decimal
    faturamento_com_pdv: Decimal
    pedidos: Decimal
    source_row_number: int

    def to_payload(self) -> dict[str, Any]:
        return {
            "filial": self.filial,
            "cod_pdv": self.cod_pdv,
            "chave": self.chave,
            "kpi": self.kpi,
            "percentual_pag_atraso": self.percentual_pag_atraso,
            "prazo_atual": self.prazo_atual,
            "cond_pag_atual": self.cond_pag_atual,
            "limite_total": _decimal_to_json(self.limite_total),
            "faturamento_com_pdv": _decimal_to_json(self.faturamento_com_pdv),
            "pedidos": _decimal_to_json(self.pedidos),
            "source_row_number": self.source_row_number,
        }


class PrazoLimiteImportService:
    def __init__(self, database_url: str, schema: str, connect_timeout_seconds: float = 3.0) -> None:
        self.database_url = database_url.strip()
        self.schema = _normalize_schema(schema)
        self.connect_timeout_seconds = max(float(connect_timeout_seconds), 1.0)

    def validate_source(self, source_path: Path) -> PrazoLimiteValidationResult:
        path = source_path.expanduser().resolve()
        errors: list[str] = []
        warnings_list: list[str] = []
        rows: list[PrazoLimiteRow] = []

        if not path.exists():
            errors.append(f"Arquivo nao encontrado: {path}")
        elif path.suffix.lower() not in EXPECTED_EXTENSION_SET:
            errors.append("Formato invalido. Use um arquivo .xlsx, .xlsm, .xls ou .csv.")
        else:
            try:
                rows = _load_prazo_limite_rows(path)
            except Exception as exc:
                errors.append(str(exc))

        if not rows and not errors:
            errors.append("Nao encontrei linhas validas na planilha de prazo e limite.")

        return PrazoLimiteValidationResult(
            dataset_name="prazo_limite",
            source_path=str(path),
            ok=not errors,
            total_rows=len(rows),
            unique_clientes=len({(row.filial, row.cod_pdv) for row in rows}),
            error_count=len(errors),
            warning_count=len(warnings_list),
            sample_errors=errors[:10],
            sample_warnings=warnings_list[:10],
        )

    def summarize_source(self, source_path: Path) -> PrazoLimiteImportSummary:
        validation = self.validate_source(source_path)
        validation.ensure_valid()
        rows = _load_prazo_limite_rows(source_path.expanduser().resolve())
        return PrazoLimiteImportSummary(
            source_path=str(source_path),
            rows=len(rows),
            unique_clientes=len({(row.filial, row.cod_pdv) for row in rows}),
            unique_filiais=len({row.filial for row in rows}),
            unique_periodos=len({row.kpi for row in rows}),
            limite_total_somado=_format_decimal_string(sum((row.limite_total for row in rows), Decimal("0"))),
            faturamento_total_somado=_format_decimal_string(sum((row.faturamento_com_pdv for row in rows), Decimal("0"))),
        )

    def import_source(self, source_path: Path, reference_date: date | None = None) -> dict[str, Any]:
        if not self.database_url:
            raise RuntimeError("REPORTS_DATABASE_URL nao configurada.")

        path = source_path.expanduser().resolve()
        validation = self.validate_source(path)
        validation.ensure_valid()
        summary = self.summarize_source(path)
        rows = _load_prazo_limite_rows(path)
        batch_date = reference_date or datetime.fromtimestamp(path.stat().st_mtime).date()
        source_hash = _sha256(path)

        with self._connect() as conn:
            self._ensure_schema(conn)
            batch_id = self._insert_batch(conn, str(path), batch_date, source_hash, len(rows))
            self._insert_snapshot_rows(conn, rows, batch_id)
            activate_import_batch(conn, self.schema, "prazo_limite", batch_id)
            self._create_latest_view(conn)
            prune_import_batches(conn, self.schema, "prazo_limite", keep_last=3)
            conn.commit()

        result = summary.to_dict()
        result.update(
            {
                "batch_id": batch_id,
                "reference_date": batch_date.isoformat(),
                "source_hash": source_hash,
                "schema": self.schema,
                "replaced_previous_batches": False,
                "published_as_active_batch": True,
            }
        )
        return result

    def refresh_latest_view(self) -> dict[str, Any]:
        if not self.database_url:
            raise RuntimeError("REPORTS_DATABASE_URL nao configurada.")

        with self._connect() as conn:
            self._ensure_schema(conn)
            active_batch_id = resolve_effective_import_batch_id(conn, self.schema, "prazo_limite", activate_if_missing=True)
            self._create_latest_view(conn)
            conn.commit()

        return {
            "ok": True,
            "schema": self.schema,
            "view": f"{self.schema}.prazo_limite_latest",
            "active_batch_id": active_batch_id,
        }

    def _ensure_schema(self, conn: psycopg.Connection[Any]) -> None:
        with conn.cursor() as cur:
            cur.execute(sql.SQL("CREATE SCHEMA IF NOT EXISTS {}").format(sql.Identifier(self.schema)))
            cur.execute(
                sql.SQL(
                    """
                    CREATE TABLE IF NOT EXISTS {}.import_batches (
                        id BIGSERIAL PRIMARY KEY,
                        dataset_name VARCHAR(80) NOT NULL,
                        source_file TEXT NOT NULL,
                        file_hash VARCHAR(64) NOT NULL,
                        reference_date DATE,
                        total_rows INTEGER NOT NULL,
                        imported_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                    """
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    """
                    CREATE TABLE IF NOT EXISTS {}.prazo_limite_snapshot (
                        batch_id BIGINT NOT NULL REFERENCES {}.import_batches(id) ON DELETE CASCADE,
                        row_number BIGINT NOT NULL,
                        filial VARCHAR(16) NOT NULL,
                        cod_pdv VARCHAR(32) NOT NULL,
                        chave TEXT NOT NULL,
                        kpi VARCHAR(40) NOT NULL,
                        percentual_pag_atraso VARCHAR(32) NOT NULL,
                        prazo_atual VARCHAR(32) NOT NULL,
                        cond_pag_atual VARCHAR(64) NOT NULL,
                        limite_total NUMERIC(18, 2) NOT NULL,
                        faturamento_com_pdv NUMERIC(18, 2) NOT NULL,
                        pedidos NUMERIC(18, 2) NOT NULL DEFAULT 0,
                        payload JSONB NOT NULL,
                        imported_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        PRIMARY KEY (batch_id, row_number)
                    )
                    """
                ).format(sql.Identifier(self.schema), sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "ALTER TABLE {}.prazo_limite_snapshot ADD COLUMN IF NOT EXISTS pedidos NUMERIC(18, 2) NOT NULL DEFAULT 0"
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS prazo_limite_snapshot_batch_filial_cod_idx ON {}.prazo_limite_snapshot (batch_id, filial, cod_pdv)"
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS prazo_limite_snapshot_batch_kpi_idx ON {}.prazo_limite_snapshot (batch_id, kpi)"
                ).format(sql.Identifier(self.schema))
            )
            cur.execute(
                sql.SQL(
                    "CREATE INDEX IF NOT EXISTS import_batches_prazo_limite_dataset_idx ON {}.import_batches (dataset_name, imported_at DESC)"
                ).format(sql.Identifier(self.schema))
            )
            ensure_dataset_state_table(conn, self.schema)

    def _insert_batch(
        self,
        conn: psycopg.Connection[Any],
        source_file: str,
        reference_date: date,
        file_hash: str,
        total_rows: int,
    ) -> int:
        query = sql.SQL(
            """
            INSERT INTO {}.import_batches (dataset_name, source_file, file_hash, reference_date, total_rows)
            VALUES ('prazo_limite', %s, %s, %s, %s)
            RETURNING id
            """
        ).format(sql.Identifier(self.schema))
        with conn.cursor(row_factory=dict_row) as cur:
            cur.execute(query, (source_file, file_hash, reference_date, total_rows))
            row = cur.fetchone()
        return int(row["id"])

    def _insert_snapshot_rows(
        self,
        conn: psycopg.Connection[Any],
        rows: list[PrazoLimiteRow],
        batch_id: int,
    ) -> None:
        query = sql.SQL(
            """
            INSERT INTO {}.prazo_limite_snapshot (
                batch_id,
                row_number,
                filial,
                cod_pdv,
                chave,
                kpi,
                percentual_pag_atraso,
                prazo_atual,
                cond_pag_atual,
                limite_total,
                faturamento_com_pdv,
                pedidos,
                payload
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
        ).format(sql.Identifier(self.schema))
        params = [
            (
                batch_id,
                index,
                row.filial,
                row.cod_pdv,
                row.chave,
                row.kpi,
                row.percentual_pag_atraso,
                row.prazo_atual,
                row.cond_pag_atual,
                row.limite_total,
                row.faturamento_com_pdv,
                row.pedidos,
                Jsonb(row.to_payload()),
            )
            for index, row in enumerate(rows, start=1)
        ]
        with conn.cursor() as cur:
            cur.executemany(query, params)

    def _create_latest_view(self, conn: psycopg.Connection[Any]) -> None:
        active_batch_id = resolve_effective_import_batch_id(conn, self.schema, "prazo_limite", activate_if_missing=True)
        where_clause = sql.SQL("s.batch_id = {}").format(sql.Literal(active_batch_id)) if active_batch_id is not None else sql.SQL("FALSE")
        query = sql.SQL(
            """
            CREATE OR REPLACE VIEW {}.prazo_limite_latest AS
            SELECT
                s.batch_id,
                s.row_number,
                s.filial,
                s.cod_pdv,
                s.chave,
                s.kpi,
                s.percentual_pag_atraso,
                s.prazo_atual,
                s.cond_pag_atual,
                s.limite_total,
                s.faturamento_com_pdv,
                s.payload,
                s.imported_at,
                b.reference_date,
                b.source_file,
                b.file_hash,
                b.imported_at AS batch_imported_at,
                s.pedidos
            FROM {}.prazo_limite_snapshot s
            JOIN {}.import_batches b ON b.id = s.batch_id
            WHERE {}
            """
        ).format(
            sql.Identifier(self.schema),
            sql.Identifier(self.schema),
            sql.Identifier(self.schema),
            where_clause,
        )
        with conn.cursor() as cur:
            cur.execute(query)

    def _connect(self) -> psycopg.Connection[Any]:
        return psycopg.connect(self.database_url, connect_timeout=self.connect_timeout_seconds)


def _load_prazo_limite_rows(path: Path) -> list[PrazoLimiteRow]:
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        return _load_prazo_limite_rows_from_workbook(path)
    return _load_prazo_limite_rows_from_csv(path)


def _load_prazo_limite_rows_from_csv(path: Path) -> list[PrazoLimiteRow]:
    text = _read_text_with_fallback(path)
    delimiter = _detect_delimiter(text)
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    if reader.fieldnames is None:
        raise ValueError("Arquivo sem cabecalho.")

    header_map: dict[str, str] = {}
    for header in reader.fieldnames:
        actual = str(header or "").strip()
        normalized = _normalize_header(actual)
        canonical = HEADER_ALIASES.get(normalized)
        if canonical and canonical not in header_map:
            header_map[canonical] = actual

        missing_headers = sorted(EXPECTED_HEADERS - set(header_map))
        if missing_headers:
            raise ValueError(f"Arquivo invalido. Colunas obrigatorias ausentes: {', '.join(missing_headers)}")

    normalized_rows = [
        {str(key or "").strip(): str(value or "").strip() for key, value in raw_row.items() if key is not None}
        for raw_row in reader
    ]
    return _build_prazo_limite_rows_from_mapping_rows(normalized_rows, header_map=header_map, row_number_offset=2)


def _load_prazo_limite_rows_from_workbook(path: Path) -> list[PrazoLimiteRow]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_values is None:
        raise ValueError("Planilha sem cabecalho.")
    header_row = [str(value or "").strip() for value in header_values]
    header_map: dict[str, str] = {}
    for actual in header_row:
        normalized = _normalize_header(actual)
        canonical = HEADER_ALIASES.get(normalized)
        if canonical and canonical not in header_map:
            header_map[canonical] = actual

    missing_headers = sorted(EXPECTED_HEADERS - set(header_map))
    if missing_headers:
        raise ValueError(f"Planilha invalida. Colunas obrigatorias ausentes: {', '.join(missing_headers)}")

    rows_as_mapping: list[dict[str, str]] = []
    for row_values in worksheet.iter_rows(min_row=2, values_only=True):
        row_map: dict[str, str] = {}
        for index, header in enumerate(header_row):
            if not header:
                continue
            value = row_values[index] if index < len(row_values) else None
            row_map[header] = _cell_to_text(value)
        rows_as_mapping.append(row_map)
    return _build_prazo_limite_rows_from_mapping_rows(rows_as_mapping, header_map=header_map, row_number_offset=2)


def _build_prazo_limite_rows_from_mapping_rows(
    raw_rows: list[dict[str, str]],
    *,
    header_map: dict[str, str],
    row_number_offset: int,
) -> list[PrazoLimiteRow]:
    rows: list[PrazoLimiteRow] = []
    for index, row in enumerate(raw_rows, start=row_number_offset):
        if not any(str(value or "").strip() for value in row.values()):
            continue
        filial = normalize_numeric_code(row.get(header_map["Filial"], ""))
        cod_pdv = normalize_numeric_code(row.get(header_map["NB"], ""))
        if not filial or not cod_pdv:
            continue
        rows.append(
            PrazoLimiteRow(
                filial=filial,
                cod_pdv=cod_pdv,
                chave=f"{filial}_{cod_pdv}",
                kpi=_clean_text(row.get(header_map["KPI"], "")) or f"Linha {index}",
                percentual_pag_atraso=_normalize_percent_text(row.get(header_map["% Pag Atraso"], "")),
                prazo_atual=_normalize_scalar_text(row.get(header_map["Prazo Atual"], "")),
                cond_pag_atual=_normalize_scalar_text(row.get(header_map["Cond Pag Atual"], "")),
                limite_total=_parse_currency_decimal(row.get(header_map["Limite Total"], "")),
                faturamento_com_pdv=_parse_currency_decimal(row.get(header_map["Faturamento com PDV"], "")),
                pedidos=_parse_decimal_value(row.get(header_map.get("Pedidos", ""), "")),
                source_row_number=index,
            )
        )
    rows.sort(key=lambda item: (_sort_numeric(item.filial), _sort_numeric(item.cod_pdv), item.source_row_number))
    return rows


def _read_text_with_fallback(path: Path) -> str:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("Nao consegui ler o arquivo com um encoding suportado.")


def _detect_delimiter(text: str) -> str:
    sample = "\n".join(text.splitlines()[:10])
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;")
        return str(dialect.delimiter or ",")
    except Exception:
        return ","


def _normalize_header(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_only = "".join(char for char in normalized if not unicodedata.combining(char))
    return "".join(char for char in ascii_only.lower() if char.isalnum())


def _normalize_schema(value: str) -> str:
    cleaned = "".join(char for char in str(value or "").strip() if char.isalnum() or char == "_")
    return cleaned or "reports"


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _normalize_scalar_text(value: Any) -> str:
    cleaned = _clean_text(value)
    if not cleaned:
        return "-"
    if cleaned.endswith(".0"):
        cleaned = cleaned[:-2]
    return cleaned


def _normalize_percent_text(value: Any) -> str:
    cleaned = _clean_text(value)
    if not cleaned:
        return "-"
    if cleaned.endswith(".0%"):
        cleaned = cleaned[:-3] + "%"
    return cleaned


def _parse_currency_decimal(value: Any) -> Decimal:
    cleaned = _clean_text(value)
    if not cleaned:
        return Decimal("0")
    normalized = cleaned.replace("R$", "").replace(" ", "")
    if "," in normalized and "." in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")
    elif "," in normalized:
        normalized = normalized.replace(",", ".")
    try:
        return Decimal(normalized)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _parse_decimal_value(value: Any) -> Decimal:
    cleaned = _clean_text(value)
    if not cleaned:
        return Decimal("0")
    normalized = cleaned.replace("%", "").replace(" ", "")
    if "," in normalized and "." in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")
    elif "," in normalized:
        normalized = normalized.replace(",", ".")
    try:
        return Decimal(normalized)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, Decimal)):
        return str(value)
    if isinstance(value, float):
        text = f"{value:.10f}".rstrip("0").rstrip(".")
        return text or "0"
    return str(value).strip()


def _format_decimal_string(value: Decimal) -> str:
    return f"{value:.2f}".replace(".", ",")


def _decimal_to_json(value: Decimal) -> str:
    return f"{value:.2f}"


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fp:
        for chunk in iter(lambda: fp.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _sort_numeric(value: str) -> tuple[int, str]:
    cleaned = normalize_numeric_code(value)
    if cleaned:
        try:
            return (0, f"{int(cleaned):09d}")
        except ValueError:
            return (1, cleaned)
    return (2, str(value or "").strip())
